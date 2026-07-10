"""커리큘럼 마스터 xlsx → 정규화 JSON 변환 파이프라인 (P0-A).

실행: PYTHONIOENCODING=utf-8 python scripts/curriculum/parse_xlsx.py
멱등(idempotent) — 같은 입력이면 항상 같은 산출물을 덮어쓴다.

입력:
    assets/level/source/한국어_4단계_통합.xlsx  (시트: 안내/0단계_자모/문법/어휘/동사)
    assets/level/source/CEFR_문장_통합.xlsx     (10,639문장 — 어휘 배분·예문의 최종 정답, 사장님 확정)
    assets/level/vocab_12levels.json            (legacy 빈도 순위 — 우선순위 점수 조인용)

산출 (assets/level/curriculum_v2/):
    grammar.json          문법 459건 (source_key=g:{교재코드}:{surface}, 레벨=교재 고정 매핑,
                          인벤토리=CEFR 12단계 통합본 기준 동기화(구표기 4 제거 + 신표기 4 추가),
                          is_core=레벨별 단원 순 상위 45개 — grammar_core_cap_v1)
    vocab.json            어휘 10,636건 (source_key=v:{surface}{접미|00},
                          레벨 배분=cefr_v1: CEFR 문장의 타깃어휘 매칭 → 문법단계+1,
                          example=해당 CEFR 문장 인라인. 매칭 실패분만 vocab_split_v1 지그재그 폴백)
    overrides.json        수동 교정 기록 (원본 유래 오류 — 사유 포함, 이 스크립트가 재생성)
    jamo.json             자모 40행 그대로 (learning_item 미적재, 프론트 화면용 정적 자산)
    survival_chunks.json  레벨 1 생존 청크 46개

추가 산출 (assets/level/level_profiles_13.json — 텍스트 자산 정리):
    profile·band·grade·stage_name·textbook 텍스트(언어학 자산)만 유지.
    (D15) grammar_count/vocab_count/grammar_scope/vocab_sample 은 산출 중단·필드 제거 —
    레벨별 문법·어휘의 단일 소스는 curriculum_v2(learning_item)다.

레벨 배분·core 선정 로직은 scripts/curriculum/assign_levels.py (순수 함수)에 위임.
근거: docs/20260709_1231_level-system-master-plan.md §2·§3
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

from assign_levels import (  # noqa: E402
    ASSIGN_RULE,
    CORE_TARGET,
    GRAMMAR_CORE_CAP,
    GRAMMAR_CORE_RULE,
    assign_vocab_levels,
    cap_grammar_core,
    compute_priority_scores,
    build_freq_percentile,
    grade_level_pair,
    rank_and_select_core,
)

_ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE_XLSX = _ROOT / "assets" / "level" / "source" / "한국어_4단계_통합.xlsx"
CEFR_XLSX = _ROOT / "assets" / "level" / "source" / "CEFR_문장_통합.xlsx"
LEGACY_FREQ_JSON = _ROOT / "assets" / "level" / "vocab_12levels.json"
OUT_DIR = _ROOT / "assets" / "level" / "curriculum_v2"
PROFILES_JSON = _ROOT / "assets" / "level" / "level_profiles_13.json"

# ---------------------------------------------------------------- 고정 매핑

# 교재 12종 고정 매핑: 교재명 → (교재코드, level_no). 미지 교재명은 즉시 에러.
TEXTBOOK_MAP: dict[str, tuple[str, int]] = {
    "Basic Korean A": ("BKA", 2), "Basic Korean B": ("BKB", 3),
    "Basic Korean C": ("BKC", 4), "Basic Korean D": ("BKD", 5),
    "Intermediate Korean A": ("IKA", 6), "Intermediate Korean B": ("IKB", 7),
    "Intermediate Korean C": ("IKC", 8), "Intermediate Korean D": ("IKD", 9),
    "Advanced Korean A": ("AKA", 10), "Advanced Korean B": ("AKB", 11),
    "Advanced Korean C": ("AKC", 12), "Advanced Korean D": ("AKD", 13),
}
GRAMMAR_ASSIGN_RULE = "textbook_v1"

# ── CEFR 12단계 통합본 (사장님 확정 최종본) ──
# 문법단계 12종 → 교재코드·앱 레벨(단계+1: A1→2 … C4→13). 기존 교재를 참고해 제작된
# 자료라 단계=교재 매핑이 성립(사장님 확인) — sync_grammar_cefr 가 인벤토리 일치를 검증한다.
CEFR_STAGES: tuple[str, ...] = (
    "A1", "A2", "A3", "A4", "B1", "B2", "B3", "B4", "C1", "C2", "C3", "C4",
)
CEFR_STAGE_TO_CODE: dict[str, str] = dict(zip(CEFR_STAGES, (
    "BKA", "BKB", "BKC", "BKD", "IKA", "IKB", "IKC", "IKD", "AKA", "AKB", "AKC", "AKD",
)))
CEFR_STAGE_TO_LEVEL: dict[str, int] = {s: i + 2 for i, s in enumerate(CEFR_STAGES)}
CEFR_ASSIGN_RULE = "cefr_v1"
# 어휘 매칭 실패 허용 상한 — 이 이상이면 표기 파서/원본 이상으로 보고 즉시 중단
CEFR_MATCH_FAIL_LIMIT = 50

# cefr_v1 재배분 후 레벨별 어휘 수 기대값 (CEFR 단계 분포 실측 — 회귀 방지 고정)
# A1 461 / A2 274 / A3 574-1 / A4 527 / B1 1163-1 / B2 494 / B3 1239 / B4 961
# / C1 1270-1 / C2 1096 / C3 1291 / C4 1289. (-1 은 여벌 문장 스킵 3건 — overrides 참조)
CEFR_LEVEL_DIST: dict[int, int] = {
    2: 461, 3: 274, 4: 573, 5: 527, 6: 1162, 7: 494,
    8: 1239, 9: 961, 10: 1269, 11: 1096, 12: 1291, 13: 1289,
}

# 품사 정규 11종 화이트리스트 (미지 토큰=파싱 에러)
POS_WHITELIST = frozenset({
    "명사", "대명사", "의존명사", "수사", "동사", "형용사",
    "부사", "관형사", "감탄사", "접사", "줄어든말",
})
POS_ALIAS = {"줄어든꼴": "줄어든말"}

_BAND_RE = re.compile(r"^([123])단계")
_GRADE_RE = re.compile(r"^([1-6])급$")
_WS_RE = re.compile(r"\s+")
_SEP_RE = re.compile(r"[/∙·‧・]")           # 표기·품사 구분자
# 토큰 → (표기, 접미 번호). 접미 해석은 직전 문자가 한글(또는 접사 표기 하이픈:
# '불-02' 류 접두사 표제어) && 숫자 1~2자리일 때만 — 'MP3' 처럼 숫자로 끝나는
# 일반(비한글) 표제어를 접미 번호로 오인하지 않게 방어. 현행 데이터 파싱 결과 불변.
_HEADWORD_RE = re.compile(r"^(.+?)(?:(?<=[가-힣-])(\d{1,2}))?$")
_EXAMPLE_LINE_RE = re.compile(r"^\d{1,2}\.\s*(.+)$")
_UNIT_NUM_RE = re.compile(r"\d+")            # 단원 번호('01과'/'1과' 혼재) 추출

# ---------------------------------------------------------------- 수동 교정 (overrides.json 원천)
# 전부 원본(국립국어원 자료) 유래 오류 — AI 정리 훼손 아님 (마스터 플랜 §3.1 전수 대조).
# 표기(cell) 비교는 공백 제거 후 수행한다. '공 적02' 류 단순 내부 공백은
# 파서 1단계(셀 전체 공백 제거)가 흡수하므로 오버라이드 대상이 아니다.

OVERRIDES: dict = {
    "notes": [
        "이 파일은 scripts/curriculum/parse_xlsx.py 가 재생성한다(수정은 파서의 OVERRIDES 상수에서).",
        "'공 적02' 류 단순 내부 공백은 파서 1단계(셀 전체 공백 제거)가 흡수 — 오버라이드 대상 아님.",
        "무접미 행('자주')과 명시적 00 행('자주00')이 같은 표기로 공존하는 동형이의어 19쌍은 "
        "오버라이드가 아니라 파서의 결정적 충돌 해소 규칙으로 처리: 무접미 행만 source_key=v:{surface}(접미 생략).",
        "cefr_sentences=CEFR_문장_통합.xlsx 문장행 교정(오타 매핑·여벌 문장 스킵), "
        "grammar_cefr=문법 인벤토리 CEFR 동기화(구표기 4 제거 + 신표기 4 추가 — 사장님 확정 최종본).",
    ],
    "grammar": [
        {
            "match": {"textbook_code": "AKD", "surface": "V-(으)ㄹ라치면"},
            "set": {
                "examples": [
                    "보고서를 낼라치면 꼭 컴퓨터가 고장 나요.",
                    "회의를 시작할라치면 전화가 울려요.",
                    "중요한 문서를 출력할라치면 프린터가 먹통이에요.",
                ],
            },
            "reason": "원본 예문 3개가 비문('내려고 하려라치면' 등 활용 오류) — 자연스러운 문장으로 교체",
        },
    ],
    "vocab": [
        {
            "match": {"grade": "1급", "cell": "명02", "guide_phrase": "병에 담다"},
            "set": {"cell": "병02"},
            "reason": "길잡이말 '병에 담다'로 보아 '병02'의 오기 확정(D9). 동일 급 '명02'(의존명사, '한 명')와 길잡이말로 구분",
        },
        {
            "match": {"grade": "2급", "cell": "대전01 명사"},
            "set": {"cell": "대전01"},
            "reason": "표기 셀에 품사('명사')가 섞여 들어간 원본 오기",
        },
        {
            "match": {"grade": "3급", "cell": "윗글0"},
            "set": {"cell": "윗글00"},
            "reason": "접미 번호 1자리 오기 → 2자리('00')로 정규화",
        },
        {
            "match": {"grade": "3급", "cell": "그제02", "seq_no": None},
            "set": {"seq_no": "grade_tail"},
            "reason": "번호 결측 — 급내 말번호(3급 최대 번호+1)를 부여",
        },
        {
            "match": {"grade": "5급", "cell": "쯧쯧01/쯧02"},
            "set": {"cell": "쯧쯧01"},
            "reason": "구분자 양쪽 표기가 다른 유일한 행('쯧쯧01' 부사 / '쯧02' 감탄사 — 이형 병기). "
                      "표기 파서 규칙 ⑤(토큰 간 표기 불일치=에러)에 걸려 대표형 '쯧쯧01'(길잡이말 '혀를 쯧쯧 차다' 일치)로 교정. "
                      "'쯧02'는 별도 표제어 아님 — 본 기록으로 원문 보존",
        },
        {
            "match": {"grade": "4급", "cell": "고소하다 01"},
            "set": {"cell": "고소하다01"},
            "reason": "접미 번호 앞 내부 공백(공백 제거 규칙과 결과 동일하나 원본 유래 오류로 명시 기록)",
        },
        {
            "match": {"grade": "6급", "cell": "자아실현 00"},
            "set": {"cell": "자아실현00"},
            "reason": "접미 번호 앞 내부 공백(공백 제거 규칙과 결과 동일하나 원본 유래 오류로 명시 기록)",
        },
    ],
    # ── CEFR_문장_통합.xlsx 문장행 교정 (타깃어휘 표기 오류·여벌 문장) ──
    # match 는 (번호, 타깃어휘 원문). set.source_key=명시 매핑 / set.skip=배분·예문 미사용.
    "cefr_sentences": [
        {
            "match": {"no": 190, "target": "명02"},
            "set": {"source_key": "v:병02"},
            "reason": "문장 '이 병에 물이 있어요.'(타깃품사 명사) — 어휘 목록 오버라이드(1급 '명02'→'병02' 오기 교정, D9)와 "
                      "동일한 원본 오류를 CEFR 이 승계. 진짜 '명02'(의존명사)는 #318 '친구 세 명이 왔어요.'가 별도 커버",
        },
        {
            "match": {"no": 3337, "target": "국05"},
            "set": {"source_key": "v:구05"},
            "reason": "문장 '이 구(區)에는 공원이 많아서 살 만해요.' — '구05'의 오타('국05'라는 표제어는 어휘 목록에 없음)",
        },
        {
            "match": {"no": 7062, "target": "쯧쯧01/쯧02"},
            "set": {"source_key": "v:쯧쯧01"},
            "reason": "어휘 목록 오버라이드(5급 이형 병기 → 대표형 '쯧쯧01')와 동일한 원본 표기를 CEFR 이 승계 — "
                      "토큰 간 표기 불일치라 표기 파서로 못 풀어 명시 매핑",
        },
        {
            "match": {"no": 8059, "target": "퇰02"},
            "set": {"source_key": "v:티02"},
            "reason": "문장 '…반가운 티가 났으련만…' — '티02'의 오타(인코딩 손상 추정 '퇰02')",
        },
        {
            "match": {"no": 750, "target": "대전"},
            "set": {"skip": True},
            "reason": "여벌 문장 — 접미 명시 행 #1815('대전01', A4)가 별도 존재. 접미 없는 중복 행은 배분·예문에 미사용",
        },
        {
            "match": {"no": 1847, "target": "구"},
            "set": {"skip": True},
            "reason": "여벌 문장 — '구05'는 #3337(오타 교정 행, B2)이 커버. 접미 없는 중복 행은 배분·예문에 미사용",
        },
        {
            "match": {"no": 5727, "target": "티"},
            "set": {"skip": True},
            "reason": "여벌 문장 — '티02'는 #8059(오타 교정 행, C2)가 커버. 접미 없는 중복 행은 배분·예문에 미사용",
        },
    ],
    # ── 문법 인벤토리 CEFR 동기화 (사장님 확정: CEFR 표기가 최종본) ──
    # 단계·단원 위치 대조 결과 구표기 4건과 신표기 4건은 단계·단원이 전혀 겹치지 않아
    # 개칭(rename) 1:1 대응이 아니라 '제거 4 + 추가 4'로 판별 — 아래에 그대로 기록.
    # (구 source_key 4건은 seed 의 stale learning_item 리포트(M4)로 표면화된다.)
    "grammar_cefr": {
        "removed": [
            {
                "match": {"textbook_code": "IKA", "surface": "어디나"},
                "reason": "CEFR B1 미수록 (구 IKA 2과 — '언제나/무엇이나'는 유지되나 '어디나'만 제외)",
            },
            {
                "match": {"textbook_code": "IKA", "surface": "V-ㄴ다고(요)/는다고(요)"},
                "reason": "CEFR B1 미수록 (구 IKA 8과 — 인접 'A-다고(요)'/'N(이)라고(요)'는 유지)",
            },
            {
                "match": {"textbook_code": "AKB", "surface": "다시 말해서 V-(으)려는지"},
                "reason": "CEFR C2 미수록 — 원본에서 '다시 말해서'(7과)와 'V-(으)려는지'(8과 별도 수록)가 병합된 손상 행",
            },
            {
                "match": {"textbook_code": "AKB", "surface": "N 길이"},
                "reason": "CEFR C2 미수록 — 원본 손상 추정 조각 표기(8과)",
            },
        ],
        # 추가분의 unit·examples 는 CEFR 문장행에서 파생, unit_title 은 같은 (교재, 단원) 기존 항목에서 차용.
        # seq_no 는 460부터 CEFR 최초 등장 번호 순으로 부여(구 459 번호는 재사용하지 않음 — 멱등·결정적).
        "added": [
            {
                "stage": "B1", "unit": "14과", "surface": "A/V-(으)면 (으)ㄹ수록",
                "reason": "CEFR B1 14과 신규 — 기존 'A/V-(으)면 A/V-(으)ㄹ수록'의 축약 반복형 변이를 별도 구조로 수록",
            },
            {
                "stage": "B3", "unit": "1과", "surface": "N(이)라면서(요)?",
                "reason": "CEFR B3 1과 신규 — 'A-다면서(요)?/V-ㄴ/는다면서(요)?' 계열의 명사 결합형",
            },
            {
                "stage": "B3", "unit": "1과", "surface": "V-(으)ㄴ 대로",
                "reason": "CEFR B3 1과 신규 — 'V-는 대로, N대로' 계열의 과거 관형형",
            },
            {
                "stage": "B4", "unit": "16과", "surface": "N(이)라기보다는",
                "reason": "CEFR B4 16과 신규 — 'A-다기보다는 / V-ㄴ/는다기보다는' 계열의 명사 결합형",
            },
        ],
    },
}

# ---------------------------------------------------------------- 생존 청크 46개 (레벨 1)

_SURVIVAL_RAW: list[tuple[str, str, str, str, str]] = [
    # (category, ko, roman, meaning_en, situation)
    ("인사", "안녕하세요?", "annyeonghaseyo?", "Hello.", "만능 인사·통화 시작"),
    ("인사", "만나서 반갑습니다", "mannaseo bangapseumnida", "Nice to meet you.", "첫 만남"),
    ("인사", "잘 지냈어요?", "jal jinaesseoyo?", "How have you been?", "아는 사이 재회"),
    ("인사", "안녕히 가세요", "annyeonghi gaseyo", "Goodbye (to one leaving).", "상대가 떠날 때"),
    ("인사", "안녕히 계세요", "annyeonghi gyeseyo", "Goodbye (I'm leaving).", "내가 떠날 때"),
    ("인사", "또 봐요", "tto bwayo", "See you again.", "가벼운 작별"),
    ("인사", "좋은 하루 보내세요", "joeun haru bonaeseyo", "Have a good day.", "작별 덧붙임"),
    ("인사", "어서 오세요", "eoseo oseyo", "Welcome.", "가게·식당에서 듣기"),
    ("감사·사과", "감사합니다", "gamsahamnida", "Thank you.", "격식 감사"),
    ("감사·사과", "고마워요", "gomawoyo", "Thanks.", "친근한 감사"),
    ("감사·사과", "아니에요", "anieyo", "Not at all.", "감사·사과 응답"),
    ("감사·사과", "죄송합니다", "joesonghamnida", "I'm sorry.", "격식 사과"),
    ("감사·사과", "미안해요", "mianhaeyo", "Sorry.", "가벼운 사과"),
    ("감사·사과", "괜찮아요", "gwaenchanayo", "It's okay.", "사과 응답·사양·안부 답"),
    ("긍정·부정·반응", "네", "ne", "Yes.·I see.", "긍정+맞장구"),
    ("긍정·부정·반응", "아니요", "aniyo", "No.", "부정"),
    ("긍정·부정·반응", "좋아요", "joayo", "Good!", "승낙·호감"),
    ("긍정·부정·반응", "맞아요", "majayo", "That's right.", "동의"),
    ("긍정·부정·반응", "알겠어요", "algesseoyo", "Got it.", "이해 확인"),
    ("긍정·부정·반응", "몰라요", "mollayo", "I don't know.", "모를 때"),
    ("긍정·부정·반응", "진짜요?", "jinjjayo?", "Really?", "놀람"),
    ("긍정·부정·반응", "맛있어요", "masisseoyo", "It's delicious.", "음식 리액션"),
    ("자기소개", "저는 ◯◯이에요", "jeoneun OOieyo", "I'm ◯◯.", "이름 말하기(슬롯)"),
    ("자기소개", "저는 ◯◯ 사람이에요", "jeoneun OO saramieyo", "I'm from ◯◯.", "국적(슬롯)"),
    ("자기소개", "이름이 뭐예요?", "ireumi mwoyeyo?", "What's your name?", "상대 이름 묻기"),
    ("자기소개", "처음 뵙겠습니다", "cheoeum boepgesseumnida", "How do you do.", "격식 첫인사"),
    ("자기소개", "잘 부탁드립니다", "jal butakdeurimnida", "Please take good care of me.", "첫 만남 마무리"),
    ("생존 요청", "이거 주세요", "igeo juseyo", "This one, please.", "주문"),
    ("생존 요청", "얼마예요?", "eolmayeyo?", "How much is it?", "가격"),
    ("생존 요청", "화장실이 어디예요?", "hwajangsiri eodiyeyo?", "Where is the bathroom?", "장소"),
    ("생존 요청", "도와주세요", "dowajuseyo", "Please help me.", "긴급"),
    ("생존 요청", "여기요!", "yeogiyo!", "Excuse me!", "점원 부르기"),
    ("생존 요청", "잠시만요", "jamsimanyo", "Just a moment.", "시간 벌기"),
    ("생존 요청", "배고파요", "baegopayo", "I'm hungry.", "상태"),
    ("생존 요청", "물 주세요", "mul juseyo", "Water, please.", "식당"),
    ("학습자 전략", "다시 말해 주세요", "dasi malhae juseyo", "Please say that again.", "못 알아들음"),
    ("학습자 전략", "천천히 말해 주세요", "cheoncheonhi malhae juseyo", "Please speak slowly.", "속도 조절"),
    ("학습자 전략", "잘 못 들었어요", "jal mot deureosseoyo", "I didn't catch that.", "청취 실패"),
    ("학습자 전략", "뭐라고요?", "mworagoyo?", "What did you say?", "되묻기"),
    ("학습자 전략", "◯◯이/가 뭐예요?", "OOi-ga mwoyeyo?", "What is ◯◯?", "단어 뜻(슬롯)"),
    ("학습자 전략", "한국어로 어떻게 말해요?", "hangugeoro eotteoke malhaeyo?", "How do you say it in Korean?", "표현 묻기"),
    ("학습자 전략", "이해 못 했어요", "ihae mot haesseoyo", "I didn't understand.", "이해 실패"),
    ("숫자", "하나·둘·셋·넷·다섯", "hana dul set net daseot", "1–5 (native)", "개수"),
    ("숫자", "여섯·일곱·여덟·아홉·열", "yeoseot ilgop yeodeol ahop yeol", "6–10 (native)", "개수"),
    ("숫자", "일·이·삼·사·오·육·칠·팔·구·십", "il i sam sa o yuk chil pal gu sip", "1–10 (Sino)", "가격·번호"),
    ("숫자", "한 개 주세요·두 개 주세요", "han gae juseyo·du gae juseyo", "One·Two, please.", "수사+단위 실전"),
]

SURVIVAL_CHUNKS: list[dict] = [
    {"no": i, "category": c, "ko": ko, "roman": r, "meaning_en": m, "situation": s}
    for i, (c, ko, r, m, s) in enumerate(_SURVIVAL_RAW, start=1)
]


class ParseError(Exception):
    """원본 데이터가 규칙을 벗어남 — 파이프라인 즉시 중단."""


def _nospace(text: str) -> str:
    return _WS_RE.sub("", text)


def _clean(value) -> str | None:
    """셀 → trim 문자열 (결측·공백만이면 None)."""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


# ---------------------------------------------------------------- 문법

def _split_examples(raw: str) -> tuple[list[str], bool]:
    """예문 컬럼을 '1./2./3.' + 줄바꿈 기준으로 분해. 실패 시 (원문 1원소, True=경고)."""
    lines = [line.strip() for line in raw.splitlines() if line.strip()]
    out: list[str] = []
    for line in lines:
        m = _EXAMPLE_LINE_RE.match(line)
        if not m:
            return [raw.strip()], True
        out.append(m.group(1).strip())
    if not out:
        return [raw.strip()], True
    return out, False


def parse_grammar(ws, overrides: list[dict]) -> list[dict]:
    """문법 시트 → 정규화 항목 리스트. 오버라이드(예문 교체) 적용 포함."""
    items: list[dict] = []
    applied = [0] * len(overrides)

    for row in ws.iter_rows(min_row=2, values_only=True):
        band_raw, seq_no, textbook, unit, unit_title, surface_raw, gtype, examples_raw, explanation, caution = row
        m = _BAND_RE.match(str(band_raw))
        if not m:
            raise ParseError(f"문법: 미지의 단계 값 {band_raw!r} (seq={seq_no})")
        band = int(m.group(1))

        if textbook not in TEXTBOOK_MAP:
            raise ParseError(f"문법: 미지의 교재명 {textbook!r} (seq={seq_no})")
        code, level_no = TEXTBOOK_MAP[textbook]

        surface = _WS_RE.sub(" ", str(surface_raw).strip())
        examples, failed = _split_examples(str(examples_raw or ""))
        if failed:
            print(f"[경고] 문법 예문 분해 실패 → 원문 1원소 유지: {code} {surface}")

        item = {
            "kind": "grammar",
            "seq_no": int(seq_no),
            "band": band,
            "textbook_code": code,
            "level_no": level_no,
            "assign_rule": GRAMMAR_ASSIGN_RULE,
            "surface": surface,
            "source_key": f"g:{code}:{surface}",
            "unit": _clean(unit),
            "unit_title": _clean(unit_title),
            "grammar_type": _nospace(gtype) if _clean(gtype) else None,
            "explanation": _clean(explanation),
            "caution": _clean(caution),
            "examples": examples,
            "is_core": True,  # 기본 True — 레벨별 상한 45('선택 문법' 분리)는 cap_grammar_core 가 적용
        }

        for i, ov in enumerate(overrides):
            match = ov["match"]
            if match["textbook_code"] == code and match["surface"] == surface:
                item.update(ov["set"])
                applied[i] += 1
        items.append(item)

    for ov, n in zip(overrides, applied):
        if n != 1:
            raise ParseError(f"문법 오버라이드 적용 {n}회(기대 1회): {ov['match']}")
    return items


# ---------------------------------------------------------------- 어휘

def _parse_headword(cell: str, ctx: str) -> tuple[str, str | None, str, bool]:
    """표기 파서: 구분자 분해 → 토큰별 (표기, 접미) → 일치 검증.

    반환 4원소: (surface, homograph_refs, source_key 접미 2자리, 무접미 여부)
    """
    tokens = [t for t in _SEP_RE.split(cell) if t]
    if not tokens:
        raise ParseError(f"어휘: 빈 표기 셀 ({ctx})")
    bases: list[str] = []
    suffixes: list[str | None] = []
    for token in tokens:
        m = _HEADWORD_RE.match(token)
        if not m:
            raise ParseError(f"어휘: 표기 토큰 파싱 실패 {token!r} ({ctx})")
        bases.append(m.group(1))
        suffixes.append(m.group(2))
    if len(set(bases)) != 1:
        raise ParseError(f"어휘: 토큰 간 표기 불일치 {bases!r} ({ctx})")
    refs = ",".join(s or "" for s in suffixes) if any(suffixes) else None
    key_suffix = suffixes[0].zfill(2) if suffixes[0] else "00"
    return bases[0], refs, key_suffix, suffixes[0] is None


def _parse_pos(raw: str, ctx: str) -> tuple[str, list[str], str]:
    """품사: 공백 제거 → 구분자 분해 → 별칭 정규화 → 화이트리스트 검증."""
    tokens = [POS_ALIAS.get(t, t) for t in _SEP_RE.split(_nospace(raw)) if t]
    if not tokens:
        raise ParseError(f"어휘: 품사 결측 ({ctx})")
    for t in tokens:
        if t not in POS_WHITELIST:
            raise ParseError(f"어휘: 미지의 품사 토큰 {t!r} (원본 {raw!r}, {ctx})")
    return tokens[0], tokens, raw


def parse_vocab(ws, overrides: list[dict]) -> list[dict]:
    """어휘 시트 → 정규화 항목 리스트. 표기 오버라이드·번호 결측 보정 적용."""
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

    # 급내 말번호(그제02 번호 결측 보정용): 급별 최대 번호
    max_seq: dict[str, int] = defaultdict(int)
    for row in raw_rows:
        if isinstance(row[2], int):
            max_seq[str(row[1])] = max(max_seq[str(row[1])], row[2])
    grade_tail_next: dict[str, int] = {g: n + 1 for g, n in max_seq.items()}

    items: list[dict] = []
    applied = [0] * len(overrides)

    for row in raw_rows:
        band_raw, grade_raw, seq_no, cell_raw, pos_raw, guide_raw = row
        grade_str = str(grade_raw).strip()
        ctx = f"{grade_str} #{seq_no} {cell_raw!r}"

        m = _BAND_RE.match(str(band_raw))
        if not m:
            raise ParseError(f"어휘: 미지의 단계 값 {band_raw!r} ({ctx})")
        band = int(m.group(1))
        gm = _GRADE_RE.match(grade_str)
        if not gm:
            raise ParseError(f"어휘: 미지의 TOPIK급 값 {grade_raw!r} ({ctx})")
        grade = int(gm.group(1))

        # ① 셀 전체 공백 제거 → ② 오버라이드 적용 (표기 비교는 공백 제거 형 기준)
        cell = _nospace(str(cell_raw))
        guide_phrase = _clean(guide_raw)
        for i, ov in enumerate(overrides):
            match = ov["match"]
            if match["grade"] != grade_str or _nospace(match["cell"]) != cell:
                continue
            if "guide_phrase" in match and match["guide_phrase"] != guide_phrase:
                continue
            if "seq_no" in match and match["seq_no"] != seq_no:
                continue
            setter = ov["set"]
            if "cell" in setter:
                cell = _nospace(setter["cell"])
            if setter.get("seq_no") == "grade_tail":
                seq_no = grade_tail_next[grade_str]
            applied[i] += 1

        if not isinstance(seq_no, int):
            raise ParseError(f"어휘: 번호 결측(오버라이드 미적용) ({ctx})")

        surface, homograph_refs, key_suffix, bare = _parse_headword(cell, ctx)
        pos_primary, pos_list, pos_original = _parse_pos(str(pos_raw), ctx)

        items.append({
            "kind": "vocab",
            "band": band,
            "topik_grade": grade,
            "seq_no": seq_no,
            "surface": surface,
            "homograph_refs": homograph_refs,
            "source_key": f"v:{surface}{key_suffix}",
            "pos_primary": pos_primary,
            "pos_list": pos_list,
            "pos_raw": pos_original,
            "guide_phrase": guide_phrase,
            "is_verb_priority": False,  # 동사 시트 조인에서 갱신
            "_raw_cell": str(cell_raw),  # 동사 시트 원문 조인용 (직렬화 전 제거)
            "_bare": bare,               # 무접미 여부 (source_key 충돌 해소용, 직렬화 전 제거)
        })

    for ov, n in zip(overrides, applied):
        if n != 1:
            raise ParseError(f"어휘 오버라이드 적용 {n}회(기대 1회): {ov['match']}")

    _resolve_bare_key_collisions(items)
    return items


def _resolve_bare_key_collisions(items: list[dict]) -> None:
    """source_key 충돌 해소 규칙 (원본 유래 동형이의어 19쌍).

    원본은 같은 표기의 동형이의어를 '무접미 행'("자주")과 '명시적 00 행'("자주00")으로
    구분하는데, 기본 규칙(무접미→'00')이 이 구분을 뭉갠다. 같은 source_key 가
    무접미 행 + 접미 명시 행 조합으로 충돌하면 **무접미 행만 접미 없이 v:{surface}** 로
    강등해 원본의 구분을 보존한다(결정적 — 행 순서 무관, 접미 조작 없음).
    그 외 패턴의 충돌은 데이터 오류이므로 그대로 에러(전역 유일 검증에서 중단).
    """
    by_key: dict[str, list[dict]] = defaultdict(list)
    for item in items:
        by_key[item["source_key"]].append(item)
    demoted = 0
    for key, group in by_key.items():
        if len(group) < 2:
            continue
        bare = [it for it in group if it["_bare"]]
        explicit = [it for it in group if not it["_bare"]]
        if len(bare) == 1 and len(explicit) == 1:
            bare[0]["source_key"] = f"v:{bare[0]['surface']}"
            demoted += 1
    if demoted:
        print(f"[정보] source_key 충돌 해소(무접미 vs 명시적 접미): {demoted}쌍 → 무접미 행은 v:{{surface}} 유지")


def flag_verb_priority(ws, vocab_items: list[dict]) -> int:
    """동사 시트를 (급, 표기 원문) 완전 일치로 어휘에 조인 → is_verb_priority. 미스 0 필수."""
    verb_keys: set[tuple[str, str]] = set()
    verb_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        verb_rows += 1
        verb_keys.add((str(row[1]).strip(), str(row[3])))
    assert len(verb_keys) == verb_rows, f"동사 시트 내 중복 표기: {verb_rows - len(verb_keys)}건"

    vocab_keys: set[tuple[str, str]] = set()
    flagged = 0
    for item in vocab_items:
        key = (f"{item['topik_grade']}급", item["_raw_cell"])
        vocab_keys.add(key)
        if key in verb_keys:
            item["is_verb_priority"] = True
            flagged += 1

    misses = verb_keys - vocab_keys
    assert not misses, f"동사→어휘 조인 미스 {len(misses)}건: {sorted(misses)[:5]}"
    return flagged


# ---------------------------------------------------------------- CEFR 12단계 통합본

def parse_cefr_sentences(ws) -> list[dict]:
    """CEFR_문장_통합.xlsx → 문장행 리스트 (번호 순, 구조 위반은 즉시 에러).

    컬럼: 번호/문법단계/문법단원/문법구조/문장/명사/동사/형용사·부사/기타어휘/
          타깃어휘/타깃품사/어휘등급(TOPIK)/문장종합레벨 (13열).
    배분·예문에 쓰는 필드(no·stage·unit·structure·sentence·target)만 보존한다.
    """
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        no, stage_raw, unit, structure, sentence = row[0], row[1], row[2], row[3], row[4]
        target = row[9]
        if not isinstance(no, int):
            raise ParseError(f"CEFR: 번호 비정수 {no!r}")
        stage = str(stage_raw).strip()
        if stage not in CEFR_STAGE_TO_LEVEL:
            raise ParseError(f"CEFR: 미지의 문법단계 {stage_raw!r} (#{no})")
        if not _clean(structure) or not _clean(sentence) or not _clean(target):
            raise ParseError(f"CEFR: 문법구조/문장/타깃어휘 결측 (#{no})")
        rows.append({
            "no": no,
            "stage": stage,
            "unit": _clean(unit),
            "structure": _WS_RE.sub(" ", str(structure).strip()),
            "sentence": str(sentence).strip(),
            "target": str(target).strip(),
        })
    rows.sort(key=lambda r: r["no"])
    if [r["no"] for r in rows] != list(range(1, len(rows) + 1)):
        raise ParseError("CEFR: 번호가 1..N 연속이 아님")
    return rows


def _cefr_target_key(target: str, vocab_keys: set[str]) -> str | None:
    """타깃어휘 표기 → vocab source_key 후보 (어휘 표기 파서 재사용).

    무접미 표기는 동형이의어 강등 키 v:{surface} 를 먼저 시도(파서 충돌 해소 규칙과 대칭),
    없으면 기본 접미 '00'. 파싱 불가(토큰 간 표기 불일치 등)는 None — 매칭 실패로 집계.
    """
    try:
        surface, _refs, key_suffix, bare = _parse_headword(_nospace(target), f"CEFR {target!r}")
    except ParseError:
        return None
    if bare and f"v:{surface}" in vocab_keys:
        return f"v:{surface}"
    return f"v:{surface}{key_suffix}"


def apply_cefr_vocab(vocab: list[dict], cefr_rows: list[dict], overrides: list[dict]) -> dict:
    """어휘 재배분(cefr_v1) + 예문 인라인 — CEFR 문장의 타깃어휘를 vocab 에 매칭.

    매칭 성공: level_no=문법단계+1, assign_rule="cefr_v1", example=그 문장(덮어씀).
    같은 어휘에 문장이 2개 이상이면 번호가 빠른 행이 이긴다(결정적) — 나머지는 dup 집계.
    실패 행은 리포트로 출력(어휘 쪽은 vocab_split_v1 지그재그 폴백 유지),
    실패가 CEFR_MATCH_FAIL_LIMIT 이상이면 파싱 에러로 중단해 원인을 파악한다.
    반환: 검증·리포트용 집계 dict.
    """
    by_key = {v["source_key"]: v for v in vocab}
    vocab_keys = set(by_key)
    applied = [0] * len(overrides)

    matched: dict[str, int] = {}          # source_key → 최초 매칭 문장 번호
    fail_rows: list[dict] = []
    skip_rows: list[dict] = []
    dup_rows: list[dict] = []

    for row in cefr_rows:  # 번호 순 (parse_cefr_sentences 가 보장)
        key: str | None = None
        skipped = False
        for i, ov in enumerate(overrides):
            m = ov["match"]
            if m["no"] != row["no"] or m["target"] != row["target"]:
                continue
            applied[i] += 1
            if ov["set"].get("skip"):
                skipped = True
            else:
                key = ov["set"]["source_key"]
            break
        if skipped:
            skip_rows.append(row)
            continue
        if key is None:
            key = _cefr_target_key(row["target"], vocab_keys)
        item = by_key.get(key) if key else None
        if item is None:
            fail_rows.append(row)
            continue
        if item["source_key"] in matched:
            dup_rows.append(row)
            continue
        matched[item["source_key"]] = row["no"]
        item["level_no"] = CEFR_STAGE_TO_LEVEL[row["stage"]]
        item["assign_rule"] = CEFR_ASSIGN_RULE
        item["example"] = row["sentence"]  # 기존 값이 있어도 CEFR 문장으로 덮어씀

    for ov, n in zip(overrides, applied):
        if n != 1:
            raise ParseError(f"CEFR 문장 오버라이드 적용 {n}회(기대 1회): {ov['match']}")

    unmatched_vocab = sorted(k for k in vocab_keys if k not in matched)

    # ── 매칭 리포트 (사장님 확인용) ──
    total = len(cefr_rows)
    print(f"  CEFR 문장 {total}행: 매칭 {len(matched)} / 스킵(여벌) {len(skip_rows)}"
          f" / 중복 {len(dup_rows)} / 실패 {len(fail_rows)}")
    print(f"  어휘 {len(vocab)}건 중 cefr_v1 {len(matched)}건"
          f" ({len(matched) / len(vocab):.4%}) — 미매칭(지그재그 폴백) {len(unmatched_vocab)}건")
    for r in skip_rows:
        print(f"    [스킵] #{r['no']} {r['stage']} 타깃 {r['target']!r} — overrides 사유 참조")
    for r in fail_rows:
        print(f"    [실패] #{r['no']} {r['stage']} 타깃 {r['target']!r} 문장 {r['sentence'][:30]!r}…")
    for k in unmatched_vocab:
        print(f"    [미매칭 어휘] {k} → vocab_split_v1 유지")

    if len(fail_rows) >= CEFR_MATCH_FAIL_LIMIT:
        raise ParseError(
            f"CEFR 타깃어휘 매칭 실패 {len(fail_rows)}건 ≥ {CEFR_MATCH_FAIL_LIMIT}"
            " — 표기 파서/원본 이상 의심, 중단"
        )
    return {
        "matched": matched,
        "fail_rows": fail_rows,
        "skip_rows": skip_rows,
        "dup_rows": dup_rows,
        "unmatched_vocab": unmatched_vocab,
    }


def sync_grammar_cefr(grammar: list[dict], cefr_rows: list[dict], spec: dict) -> None:
    """문법 인벤토리 CEFR 동기화 — 구표기 4 제거 + 신표기 4 추가 (grammar 리스트 in-place).

    추가 항목은 CEFR 문장행에서 단계·단원·예문을 파생하고, unit_title 은 같은
    (교재, 단원) 기존 항목에서 차용한다. 마지막에 단계별 구조 집합이 교재별 surface
    집합과 완전히 일치하는지 전수 대조(불일치=에러 — CEFR 이 최종본임을 기계적으로 보증).
    """
    # ① 제거 (정확히 1건씩)
    for op in spec["removed"]:
        m = op["match"]
        hits = [g for g in grammar
                if g["textbook_code"] == m["textbook_code"] and g["surface"] == m["surface"]]
        if len(hits) != 1:
            raise ParseError(f"문법 CEFR 제거 대상 {len(hits)}건(기대 1건): {m}")
        grammar.remove(hits[0])

    # ② 추가 재료: 구조(공백 정규화)별 문장행, (교재, 단원 숫자)별 unit_title
    struct_rows: dict[str, list[dict]] = defaultdict(list)
    for r in cefr_rows:
        struct_rows[_nospace(r["structure"])].append(r)
    unit_titles: dict[tuple[str, int], str | None] = {}
    for g in sorted(grammar, key=lambda g: -g["seq_no"]):  # 최소 seq_no 항목이 최종 승자
        um = _UNIT_NUM_RE.search(str(g.get("unit") or ""))
        if um:
            unit_titles[(g["textbook_code"], int(um.group()))] = g["unit_title"]

    next_seq = 460  # 구 1~459 번호는 재사용하지 않는다(결정적·멱등)
    for op in spec["added"]:
        surface = _WS_RE.sub(" ", op["surface"].strip())
        rows = sorted(struct_rows.get(_nospace(surface), []), key=lambda r: r["no"])
        if not rows:
            raise ParseError(f"문법 CEFR 추가 대상이 CEFR 문장에 없음: {surface!r}")
        stages = {r["stage"] for r in rows}
        units = {r["unit"] for r in rows}
        if stages != {op["stage"]} or units != {op["unit"]}:
            raise ParseError(
                f"문법 CEFR 추가 {surface!r} 위치 불일치: 기대 {op['stage']} {op['unit']}"
                f" / 실측 {sorted(stages)} {sorted(units)}"
            )
        code = CEFR_STAGE_TO_CODE[op["stage"]]
        level_no = CEFR_STAGE_TO_LEVEL[op["stage"]]
        um = _UNIT_NUM_RE.search(str(op["unit"]))
        grammar.append({
            "kind": "grammar",
            "seq_no": next_seq,
            "band": 1 if level_no <= 5 else (2 if level_no <= 9 else 3),
            "textbook_code": code,
            "level_no": level_no,
            "assign_rule": CEFR_ASSIGN_RULE,
            "surface": surface,
            "source_key": f"g:{code}:{surface}",
            "unit": op["unit"],
            "unit_title": unit_titles.get((code, int(um.group()))) if um else None,
            "grammar_type": None,
            "explanation": None,
            "caution": None,
            "examples": [r["sentence"] for r in rows],
            "is_core": True,  # 상한 45 는 cap_grammar_core 가 적용
        })
        next_seq += 1

    # ③ 전수 대조: 단계별 CEFR 구조 집합 == 교재별 문법 surface 집합 (공백 정규화)
    stage_sets: dict[str, set[str]] = defaultdict(set)
    for r in cefr_rows:
        stage_sets[r["stage"]].add(_nospace(r["structure"]))
    code_sets: dict[str, set[str]] = defaultdict(set)
    for g in grammar:
        code_sets[g["textbook_code"]].add(_nospace(g["surface"]))
    for stage, code in CEFR_STAGE_TO_CODE.items():
        if stage_sets[stage] != code_sets[code]:
            only_cefr = sorted(stage_sets[stage] - code_sets[code])
            only_old = sorted(code_sets[code] - stage_sets[stage])
            raise ParseError(
                f"문법 인벤토리 불일치 {stage}↔{code}: CEFR에만 {only_cefr} / 교재에만 {only_old}"
            )


# ---------------------------------------------------------------- 자모

def parse_jamo(ws) -> list[dict]:
    """0단계_자모 시트 40행을 헤더 그대로 보존('No'만 'no'로)."""
    rows = list(ws.iter_rows(values_only=True))
    header = ["no" if h == "No" else str(h) for h in rows[0]]
    return [dict(zip(header, row)) for row in rows[1:] if row[0] is not None]


# ---------------------------------------------------------------- 레벨 프로파일 정리

# (D15) 프로파일에서 제거된 파생 필드 — 단일 소스는 curriculum_v2(learning_item).
PROFILE_REMOVED_FIELDS = ("grammar_count", "vocab_count", "grammar_scope", "vocab_sample")

_PROFILES_COMMENT = (
    "normalcall 레벨 마스터 시드 소스(13단계). DB level 테이블 시드 + 프롬프트 [학습자 수준] 슬롯 주입용. "
    "레벨 1=생존 회화(assets/level/curriculum_v2/survival_chunks.json 46청크, 레벨테스트 배정 전용). "
    "profile·band·grade·stage_name·textbook 텍스트는 수기 언어학 자산 — 이 파일에서 직접 관리(파서가 유지). "
    "(D15) grammar_count/vocab_count/grammar_scope/vocab_sample 파생 필드는 폐지 — "
    "레벨별 문법·어휘의 단일 소스는 curriculum_v2(learning_item)."
)


def rebuild_level_profiles() -> dict:
    """level_profiles_13.json 을 텍스트 자산(이름류)만으로 재구성한다(D15).

    profile·band·grade·stage_name·textbook 는 기존 파일 값 유지(언어학 자산).
    파생 필드(grammar_count/vocab_count/grammar_scope/vocab_sample)는 제거 —
    curriculum_v2 가 단일 소스라 이중 캐시였다. 재실행 멱등 — 같은 입력이면 같은 파일.
    """
    data = json.loads(PROFILES_JSON.read_text(encoding="utf-8"))
    data["_comment"] = _PROFILES_COMMENT
    for entry in data["levels"]:
        for field in PROFILE_REMOVED_FIELDS:
            entry.pop(field, None)
    return data


# ---------------------------------------------------------------- 검증

def validate(
    grammar: list[dict],
    vocab: list[dict],
    jamo: list[dict],
    verb_flagged: int,
    core_counts: dict[int, int],
    grammar_core_counts: dict[int, int],
    profiles: dict,
    cefr_report: dict,
) -> None:
    """완료 기준 전체 검증 — 하나라도 어긋나면 AssertionError 로 중단."""
    checks: list[tuple[str, bool]] = []

    def check(name: str, ok: bool) -> None:
        checks.append((name, ok))

    check("문법 459건", len(grammar) == 459)
    gkeys = Counter(g["source_key"] for g in grammar)
    check("문법 source_key 전역 유일", all(c == 1 for c in gkeys.values()))
    seqs = sorted(g["seq_no"] for g in grammar)
    check("문법 seq_no 전역 유일(1~463 내 459개, CEFR 추가분 460~463)",
          len(set(seqs)) == 459 and seqs[0] >= 1 and seqs[-1] == 463)

    # 문법 인벤토리 CEFR 동기화: 신표기 4 존재 · 구표기 4 부재 (공백 정규화 비교)
    g_nospace = {_nospace(g["surface"]) for g in grammar}
    new_surfaces = [_nospace(op["surface"]) for op in OVERRIDES["grammar_cefr"]["added"]]
    old_surfaces = [_nospace(op["match"]["surface"]) for op in OVERRIDES["grammar_cefr"]["removed"]]
    check("문법 CEFR 신표기 4건 존재", all(s in g_nospace for s in new_surfaces))
    check("문법 CEFR 구표기 4건 부재", all(s not in g_nospace for s in old_surfaces))

    lachimyeon = next(g for g in grammar if g["surface"] == "V-(으)ㄹ라치면")
    check("예문 교정 1건(ㄹ라치면) 적용", lachimyeon["examples"] == [
        "보고서를 낼라치면 꼭 컴퓨터가 고장 나요.",
        "회의를 시작할라치면 전화가 울려요.",
        "중요한 문서를 출력할라치면 프린터가 먹통이에요.",
    ])

    check("어휘 10,636건", len(vocab) == 10636)
    vkeys = Counter(v["source_key"] for v in vocab)
    dup = [k for k, c in vkeys.items() if c > 1]
    check(f"어휘 source_key 전역 유일 (중복 {len(dup)}건)", not dup)
    if dup:
        print("  중복 키 예시:", dup[:10])
    check("is_verb_priority 1,468건", verb_flagged == 1468 and sum(v["is_verb_priority"] for v in vocab) == 1468)
    check("어휘 오버라이드 반영(병02/대전01/윗글00/그제02/고소하다01/자아실현00/쯧쯧01)", all([
        any(v["source_key"] == "v:쯧쯧01" for v in vocab),
        any(v["source_key"] == "v:병02" and v["guide_phrase"] == "병에 담다" for v in vocab),
        any(v["source_key"] == "v:대전01" for v in vocab),
        any(v["source_key"] == "v:윗글00" for v in vocab),
        any(v["source_key"] == "v:그제02" and v["seq_no"] == 1656 for v in vocab),  # 3급 말번호
        any(v["source_key"] == "v:고소하다01" for v in vocab),
        any(v["source_key"] == "v:자아실현00" for v in vocab),
    ]))
    check("전 항목 level_no·assign_rule·priority_rank 채움", all(
        v.get("level_no") and v.get("assign_rule") in (CEFR_ASSIGN_RULE, ASSIGN_RULE)
        and v.get("priority_rank") for v in vocab
    ))

    # cefr_v1 재배분: 매칭률 ≥99.5% + 매칭분 example 전건 채움 + 레벨 분포 = CEFR 단계 분포
    cefr_items = [v for v in vocab if v["assign_rule"] == CEFR_ASSIGN_RULE]
    match_rate = len(cefr_items) / len(vocab)
    check(f"cefr_v1 매칭률 {match_rate:.4%} ≥ 99.5%", match_rate >= 0.995)
    check("cefr_v1 예문 채움율 100%", all(v.get("example") for v in cefr_items))
    cefr_dist = dict(Counter(v["level_no"] for v in cefr_items))
    check("cefr_v1 레벨 분포 = CEFR 단계 분포", cefr_dist == CEFR_LEVEL_DIST)
    check("cefr_v1 매칭 집계 정합(리포트 ↔ 항목)", len(cefr_report["matched"]) == len(cefr_items))

    # 레벨-급 정합: 각 어휘의 level_no 가 소속 TOPIK급의 레벨쌍 안 (cefr/폴백 공통)
    check("전 어휘 level_no ∈ TOPIK급 레벨쌍", all(
        v["level_no"] in grade_level_pair(v["topik_grade"]) for v in vocab
    ))

    # core 레벨별 목표치
    for level_no in sorted(CORE_TARGET):
        got = core_counts.get(level_no, 0)
        check(f"L{level_no} core {got}/{CORE_TARGET[level_no]}", got == CORE_TARGET[level_no])
    check("core 금지 품사 미포함", not any(
        v["is_core"] and ("접사" in v["pos_list"] or "줄어든말" in v["pos_list"] or v["pos_list"] == ["의존명사"])
        for v in vocab
    ))

    # 문법 core 상한 45 (grammar_core_cap_v1)
    grammar_per_level = Counter(g["level_no"] for g in grammar)
    for level_no in sorted(grammar_per_level):
        expect = min(grammar_per_level[level_no], GRAMMAR_CORE_CAP)
        got = grammar_core_counts.get(level_no, 0)
        check(f"L{level_no} 문법 core {got}/min({grammar_per_level[level_no]},{GRAMMAR_CORE_CAP})", got == expect)
    check("문법 assign_rule 에 core cap 버전 기록", all(
        GRAMMAR_CORE_RULE in g["assign_rule"] for g in grammar
    ))

    # 레벨 프로파일(1~13) — 텍스트 자산 보존 + 파생 필드 부재(D15)
    profile_map = {e["level_no"]: e for e in profiles["levels"]}
    check("프로파일 파생 필드 부재(D15 — count/scope/sample 제거)", all(
        f not in profile_map[lv] for lv in range(1, 14) for f in PROFILE_REMOVED_FIELDS
    ))
    check("프로파일 텍스트 자산(profile·stage_name) 보존", all(
        profile_map[lv].get("profile") and profile_map[lv].get("stage_name")
        for lv in range(1, 14)
    ))

    check("자모 40행", len(jamo) == 40)
    check("생존 청크 46개", len(SURVIVAL_CHUNKS) == 46 and [c["no"] for c in SURVIVAL_CHUNKS] == list(range(1, 47)))

    failed = [name for name, ok in checks if not ok]
    for name, ok in checks:
        print(f"  {'OK ' if ok else 'FAIL'} {name}")
    assert not failed, f"검증 실패 {len(failed)}건: {failed}"


# ---------------------------------------------------------------- 저장·메인

def _dump(path: Path, payload) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"생성: {path.relative_to(_ROOT)}")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True)
    cefr_wb = openpyxl.load_workbook(CEFR_XLSX, read_only=True)

    print("[1/7] 파싱 (교재 통합본 + CEFR 12단계 통합본)")
    grammar = parse_grammar(wb["문법"], OVERRIDES["grammar"])
    vocab = parse_vocab(wb["어휘"], OVERRIDES["vocab"])
    verb_flagged = flag_verb_priority(wb["동사"], vocab)
    jamo = parse_jamo(wb["0단계_자모"])
    cefr_rows = parse_cefr_sentences(cefr_wb[cefr_wb.sheetnames[0]])

    print("[2/7] 문법 인벤토리 CEFR 동기화 (구표기 4 제거 + 신표기 4 추가 + 전수 대조)")
    sync_grammar_cefr(grammar, cefr_rows, OVERRIDES["grammar_cefr"])

    print("[3/7] 우선순위 점수 (빈도 40 + 품사 35 + 교재등장 25)")
    legacy = json.loads(LEGACY_FREQ_JSON.read_text(encoding="utf-8"))
    compute_priority_scores(vocab, grammar, build_freq_percentile(legacy))

    print("[4/7] 레벨 배분(cefr_v1 — 지그재그는 매칭 실패분 폴백) + core 선정")
    assign_vocab_levels(vocab)  # vocab_split_v1 베이스라인(폴백)
    cefr_report = apply_cefr_vocab(vocab, cefr_rows, OVERRIDES["cefr_sentences"])
    core_counts = rank_and_select_core(vocab)

    print(f"[5/7] 문법 core 상한 {GRAMMAR_CORE_CAP} ({GRAMMAR_CORE_RULE})")
    grammar_core_counts = cap_grammar_core(grammar)

    print("[6/7] 저장 + 레벨 프로파일 정리(텍스트 자산만 — D15)")
    for item in vocab:
        item.pop("_raw_cell", None)
        item.pop("_bare", None)
    vocab.sort(key=lambda v: (v["topik_grade"], v["seq_no"]))
    meta = {"source": "assets/level/source/한국어_4단계_통합.xlsx",
            "cefr_source": "assets/level/source/CEFR_문장_통합.xlsx",
            "generated_by": "scripts/curriculum/parse_xlsx.py"}
    grammar_rule = f"{GRAMMAR_ASSIGN_RULE}+{CEFR_ASSIGN_RULE}+{GRAMMAR_CORE_RULE}"
    vocab_rule = f"{CEFR_ASSIGN_RULE}({ASSIGN_RULE} fallback)"
    _dump(OUT_DIR / "grammar.json", {**meta, "assign_rule": grammar_rule, "count": len(grammar), "items": grammar})
    _dump(OUT_DIR / "vocab.json", {**meta, "assign_rule": vocab_rule, "count": len(vocab), "items": vocab})
    _dump(OUT_DIR / "overrides.json", OVERRIDES)
    _dump(OUT_DIR / "jamo.json", {**meta, "count": len(jamo), "items": jamo})
    _dump(OUT_DIR / "survival_chunks.json", {"level_no": 1, "count": len(SURVIVAL_CHUNKS), "items": SURVIVAL_CHUNKS})
    profiles = rebuild_level_profiles()
    _dump(PROFILES_JSON, profiles)

    print("[7/7] 검증")
    validate(grammar, vocab, jamo, verb_flagged, core_counts, grammar_core_counts, profiles, cefr_report)
    print("검증 통과 — curriculum_v2 산출 완료")


if __name__ == "__main__":
    main()
