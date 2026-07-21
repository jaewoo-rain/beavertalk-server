"""일본어 커리큘럼 raw → 정규화 JSON 변환 (멀티랭귀지 T4b · 일본어 수직 관통).

실행: PYTHONIOENCODING=utf-8 conda run -n beavertalk-server python scripts/curriculum/parse_jp.py
멱등 — 같은 입력이면 항상 같은 산출물을 덮어쓴다.

한국어(parse_xlsx.py)와 **같은 중간포맷**을 산출해 scripts/seed.py 가 그대로 적재하게 한다.
다만 원본 데이터 형태가 전혀 달라(한국어 xlsx 시트 vs 일본어 JSON+CEFR문장통합) 파서는 별도다.
한국어 cefr_v1 과 개념은 동일: **CEFR 문장통합본의 타깃어휘·문법단계가 어휘 레벨의 정답**이다.

입력 (level/05.다른 언어 CEFR/jp_cefr_rebuild/):
    JP_CEFR_문장_통합.xlsx  9,018문장 (번호/문법단계/문법단원/문법구조/문장/…/타깃어휘/타깃품사/어휘등급JEV)
                            → 어휘 레벨 배분·예문의 정답. 타깃어휘 dedup(최저 CEFR 단계 승).
    jp_grammar_12.json      문법 691건 (문법단계/문법구조/예문/문법단원/토픽/분류)
    jp_vocab_by_lv.json     표제어 17,206 (표기→읽기[가나]·한국어뜻 조인 소스)

산출 (assets/level/curriculum_v2_ja/):
    grammar.json          문법 (source_key=g:ja:{surface}, level_no=문법단계+1[A1→2..C4→13],
                          is_core=레벨별 단원 순 상위 45 — 한국어 grammar_core_cap 과 동수)
    vocab.json            어휘 (source_key=v:ja:{surface}, level_no=문법단계+1, example=CEFR 문장,
                          reading=가나, meanings={"ko":한국어뜻}, is_core=레벨별 상위 100)
추가 산출 (assets/level/level_profiles_ja.json):
    레벨 1(생존)~13(C4) 프로파일 골격. profile 본문은 T5(콘텐츠 저작)에서 채운다 — 여기선 placeholder.

레벨 구조(플랜 결정 line 99): level_no 전역 고정 — 1≡생존, 2~13≡A1~C4(한국어와 동일 축).
생존청크(level 1)·레벨 프로파일 본문·레벨테스트 앵커는 T5 저작 대상(여기 미포함).
"""

from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parent.parent.parent
_SRC = _ROOT / "level" / "05.다른 언어 CEFR" / "jp_cefr_rebuild"
SENT_XLSX = _SRC / "JP_CEFR_문장_통합.xlsx"
GRAMMAR_JSON = _SRC / "jp_grammar_12.json"
VOCAB_BY_LV_JSON = _SRC / "jp_vocab_by_lv.json"

OUT_DIR = _ROOT / "assets" / "level" / "curriculum_v2_ja"
PROFILES_JSON = _ROOT / "assets" / "level" / "level_profiles_ja.json"

# CEFR 12단계 → 앱 레벨(단계+1: A1→2 … C4→13). 한국어와 동일 축(1=생존).
CEFR_STAGES: tuple[str, ...] = (
    "A1", "A2", "A3", "A4", "B1", "B2", "B3", "B4", "C1", "C2", "C3", "C4",
)
STAGE_TO_LEVEL: dict[str, int] = {s: i + 2 for i, s in enumerate(CEFR_STAGES)}
STAGE_IDX: dict[str, int] = {s: i for i, s in enumerate(CEFR_STAGES)}

ASSIGN_RULE = "cefr_ja_v1"
GRAMMAR_CORE_CAP = 45   # 레벨별 문법 core 상한(한국어 grammar_core_cap_v1 과 동수 — 승급 게이트 크기 일치)
VOCAB_CORE_CAP = 100    # 레벨별 어휘 core 상한(대략 한국어 초급 목표치)


class ParseError(Exception):
    """원본이 규칙을 벗어남 — 즉시 중단."""


def _band(level_no: int) -> int:
    """레벨 → 밴드(2~5 초급=1 / 6~9 중급=2 / 10~13 고급=3). 한국어와 동일 경계."""
    return 1 if level_no <= 5 else (2 if level_no <= 9 else 3)


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ---------------------------------------------------------------- 읽기·한국어뜻 조인 소스

def build_reading_gloss() -> tuple[dict[str, str], dict[str, str]]:
    """jp_vocab_by_lv.json → (표기→읽기[가나], 표기→한국어뜻). 첫 등장 우선."""
    data = json.loads(VOCAB_BY_LV_JSON.read_text(encoding="utf-8"))
    reading: dict[str, str] = {}
    gloss: dict[str, str] = {}
    for items in data.values():
        for it in items:
            s = _clean(it.get("표기"))
            if not s:
                continue
            if s not in reading:
                r = _clean(it.get("읽기"))
                if r:
                    reading[s] = r
                g = _clean(it.get("한국어뜻"))
                if g:
                    gloss[s] = g
    return reading, gloss


# ---------------------------------------------------------------- 어휘 (CEFR 문장통합본)

def parse_vocab(reading: dict[str, str], gloss: dict[str, str]) -> list[dict]:
    """CEFR 문장통합본 타깃어휘 → 어휘 항목. 최저 CEFR 단계가 이기고, 그 문장을 예문으로.

    한국어 cefr_v1 과 동형: (surface) 별 최초(최저단계) 출현이 레벨·예문을 확정한다.
    """
    wb = openpyxl.load_workbook(SENT_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    # surface → 확정 레코드 (최저 단계 우선; 동단계면 번호 빠른 문장)
    chosen: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        no, stage, unit, struct, sentence = row[0], row[1], row[2], row[3], row[4]
        target, tpos, jev = row[9], row[10], row[11]
        stage = _clean(stage)
        surface = _clean(target)
        if not stage or not surface:
            continue
        if stage not in STAGE_TO_LEVEL:
            raise ParseError(f"어휘: 미지의 문법단계 {stage!r} (#{no})")
        si = STAGE_IDX[stage]
        prev = chosen.get(surface)
        if prev is not None and (si, no) >= (prev["_si"], prev["_no"]):
            continue
        chosen[surface] = {
            "surface": surface, "stage": stage, "_si": si, "_no": no,
            "sentence": _clean(sentence), "pos": _clean(tpos),
            "jev": _jev_int(jev),
        }
    wb.close()

    items: list[dict] = []
    for c in chosen.values():
        level_no = STAGE_TO_LEVEL[c["stage"]]
        surface = c["surface"]
        pos = c["pos"]
        meanings = {"ko": gloss[surface]} if surface in gloss else None
        items.append({
            "kind": "vocab",
            "source_key": f"v:ja:{surface}",
            "band": _band(level_no),
            "topik_grade": c["jev"],          # JEV 급(1~6) — 참조용(일본어엔 TOPIK 없음)
            "level_no": level_no,
            "assign_rule": ASSIGN_RULE,
            "surface": surface,
            "reading": reading.get(surface),  # 가나 (없으면 None)
            "pos_primary": pos,
            "pos_list": [pos] if pos else None,
            "pos_raw": pos,
            "is_verb_priority": False,
            "is_core": False,                 # rank_and_cap 에서 레벨별 상위 100 True
            "example": c["sentence"],
            "meanings": meanings,
            "_si": c["_si"], "_no": c["_no"],  # 정렬용(직렬화 전 제거)
        })
    _rank_and_cap(items, VOCAB_CORE_CAP)
    return items


def _jev_int(raw) -> int | None:
    """'JEV3' → 3. 결측/이상은 None."""
    s = _clean(raw)
    if not s:
        return None
    digits = "".join(ch for ch in s if ch.isdigit())
    return int(digits) if digits else None


# ---------------------------------------------------------------- 문법 (jp_grammar_12.json)

def parse_grammar() -> list[dict]:
    """jp_grammar_12.json → 문법 항목. surface 중복은 최저 단계가 승(dedup)."""
    data = json.loads(GRAMMAR_JSON.read_text(encoding="utf-8"))
    chosen: dict[str, dict] = {}
    for i, e in enumerate(data):
        stage = _clean(e.get("문법단계"))
        surface = _clean(e.get("문법구조"))
        if not stage or not surface:
            continue
        if stage not in STAGE_TO_LEVEL:
            raise ParseError(f"문법: 미지의 단계 {stage!r} (surface={surface!r})")
        si = STAGE_IDX[stage]
        seq = e.get("번호") if isinstance(e.get("번호"), int) else i
        prev = chosen.get(surface)
        if prev is not None and (si, seq) >= (prev["_si"], prev["_seq"]):
            continue
        chosen[surface] = {
            "surface": surface, "stage": stage, "_si": si, "_seq": seq,
            "unit": _clean(e.get("문법단원")), "unit_title": _clean(e.get("토픽")),
            "grammar_type": _clean(e.get("분류")),
            "example": _clean(e.get("예문")),
        }
    items: list[dict] = []
    for c in chosen.values():
        level_no = STAGE_TO_LEVEL[c["stage"]]
        items.append({
            "kind": "grammar",
            "source_key": f"g:ja:{c['surface']}",
            "band": _band(level_no),
            "textbook_code": f"JP-{c['stage']}",
            "level_no": level_no,
            "assign_rule": ASSIGN_RULE,
            "surface": c["surface"],
            "reading": None,
            "unit": c["unit"],
            "unit_title": c["unit_title"],
            "grammar_type": c["grammar_type"],
            "examples": [c["example"]] if c["example"] else None,
            "explanation": None,
            "caution": None,
            "is_core": False,   # rank_and_cap 에서 레벨별 상위 45 True
            "_si": c["_si"], "_seq": c["_seq"],
        })
    _rank_and_cap(items, GRAMMAR_CORE_CAP)
    return items


# ---------------------------------------------------------------- core 선정·순위

def _rank_and_cap(items: list[dict], cap: int) -> None:
    """레벨 내 (단계, 원본 순서)로 정렬 → priority_rank 부여 + 상위 cap 개 is_core=True."""
    by_level: dict[int, list[dict]] = defaultdict(list)
    for it in items:
        by_level[it["level_no"]].append(it)
    for level_no, group in by_level.items():
        group.sort(key=lambda x: (x["_si"], x.get("_no", x.get("_seq", 0))))
        for rank, it in enumerate(group, start=1):
            it["priority_rank"] = rank
            it["seq_no"] = rank
            it["is_core"] = rank <= cap


# ---------------------------------------------------------------- 레벨 프로파일 (T5 저작)
# profile 본문은 실제 시드된 일본어 문법 앵커(레벨별 문법구조)에 근거해 저작한 언어학 자산이다.
# 프롬프트 [학습자 수준] 슬롯에 주입되어 비버가 "이 레벨 학습자가 어떤 문법·문장으로 말하는가"를
# 파악하게 한다. 한국어(level_profiles_13.json)와 같은 밀도·문체 — 한국어 메타로 일본어 형태를 인용.
# 도그푸딩 대상이 한국인(L1=ko)이라 메타·본문은 한국어로 쓴다.
STAGE_NAME: dict[int, str] = {
    1: "생존 회화",
    2: "초급 1 (A1)", 3: "초급 2 (A2)", 4: "초급 3 (A3)", 5: "초급 4 (A4)",
    6: "중급 1 (B1)", 7: "중급 2 (B2)", 8: "중급 3 (B3)", 9: "중급 4 (B4)",
    10: "고급 1 (C1)", 11: "고급 2 (C2)", 12: "고급 3 (C3)", 13: "고급 4 (C4)",
}

PROFILE_TEXT: dict[int, str] = {
    1: "인사·숫자·자기소개·정형 표현 46개를 통째로 익히는 생존 단계. 문법 규칙 없이 "
       "「おはよう」「ありがとう」「すみません」「〜をください」 같은 덩어리 표현을 상황별로 말한다. "
       "조사·활용을 분석하지 않고 정형구 그대로 사용한다.",
    2: "가장 기초. 「〜は〜です」 명사문과 조사 は・も・と・の・を・で, 「〜が好き」, "
       "기본 동사 「〜を〜ます」, 「〜に住んでいる」까지만. です・ます체 짧은 단문(5~10모라)으로 말한다. "
       "과거형·연결형·복잡한 활용은 아직 쓰지 않는다.",
    3: "존재·이동·권유. 「〜がある/いる」, 「〜に行く」, 권유 「〜ませんか/ましょう」, "
       "범위 「〜から〜まで」, 이유 「〜から」, 형용사 수식 「形容詞＋N」, 희망 「〜が欲しい」. "
       "い·な형용사 기본 활용이 시작된다. 여전히 です・ます체 단문 중심, 복문·과거는 제한적.",
    4: "진행·과거·변화. 「〜ている」 진행/상태, 과거 「〜でした/かったです/ました」, "
       "시간 「〜とき」, 변화 「〜になる/くなる」, 명사화 「〜こと」. 시제를 현재·과거로 구분해 말한다. "
       "아직 조건·수동·경어는 쓰지 않는다.",
    5: "의도·시도·비교. 희망 「〜たい/たくない」, 시도 「〜てみる」, 방향 「〜ていく/くる」, "
       "정중 의뢰 「〜てくださいませんか」, 비교 「〜と〜とどちらが」, 방법 「〜かた」. "
       "て형을 활용한 복합 표현이 등장하고 문장이 2절로 길어지기 시작한다.",
    6: "인용·이유·조건·금지. 「〜という/という意味」, 이유 「〜ので」, 조건 「〜なら」, "
       "금지 「〜てはいけない」, 순서 「〜てから」, 양태 「〜そうだ」. 복문을 이유·조건으로 연결한다. "
       "반말(だ체)과 です・ます체를 상황에 맞게 고르기 시작한다.",
    7: "목적·난이도·완료. 목적 「〜ように」, 나열 「〜たり」, 완료·후회 「〜てしまう」, "
       "난이 「〜にくい/やすい」, 비교 「〜のほうが」, 역접 「〜ても」. 여러 절을 목적·양보로 엮은 문장을 만든다. "
       "화제 전환·부연이 자연스러워진다.",
    8: "의무·전문·조건 확장. 의무 「〜なければならない」, 전문 「〜って言っていました」, "
       "「実は〜んです」 설명, 조건 「〜ば/なければ」, 「〜けど〜から」 복합 접속. "
       "자기 의견을 이유와 함께 전개하고 だ체를 자연스럽게 구사한다.",
    9: "경어·추량·동작 국면. 존경어 「尊敬の動詞/お〜になる」, 추량 「〜みたいだ/でしょうか」, "
       "동작 개시 「〜はじめる」, 준비 「〜ておく」, 시간 「〜たあと」, 連用中止. "
       "겸양·존경을 청자에 따라 조절하고 격식·비격식을 구분한다.",
    10: "구어·수동·완곡. 구어 「〜っていうか/っていうより」, 수동·자발 「〜(ら)れる」, "
        "완곡 부정 「〜ないこともない」, 확신 「〜に違いない」, 「〜たばかりだ」, 「〜しかない」. "
        "뉘앙스를 섬세하게 조절하는 구어체로 추상적 화제도 소화한다.",
    11: "격식 접속·서면체. 「〜にあたって/に先立って/をきっかけに/ゆえに/ことから」, "
        "「〜に伴って」, 「〜に応じて」, 「〜あまり」 등 문어·격식 접속 표현으로 논리를 전개한다. "
        "뉴스·논설 수준의 문장 구조를 구사한다.",
    12: "고급 격식 관용. 「〜に難くない/を余儀なくされる/を禁じ得ない/ずにはおかない/"
        "をものともせず/ならでは/たる者」 등 문어 관용 표현. 감정·평가를 격식체 관용구로 압축한다. "
        "원어민 서면 수준.",
    13: "최상급 문어·수사. 「〜にもほどがある/それまでだ/に定評がある/そのもの/かたがた/"
        "んとする/に相違ない」 등 고급 수사·강조 표현. 뉘앙스·강조·반어를 자유자재로 구사한다. "
        "문학·전문 담화 수준의 완성된 일본어.",
}


# ---------------------------------------------------------------- 생존청크 46 (레벨 1, T5 저작)
# 한국어 생존청크 46슬롯(parse_xlsx._SURVIVAL_RAW)과 같은 기능 영역을 자연스러운 일본어로 저작.
# 정형 표현을 덩어리째 익히는 레벨 1 — 문법 분석 없이 상황별로 통째 사용.
# (category, surface[일본어], reading[가나], roman[로마자], meaning_en, meaning_ko, situation[한국어]).
_SURVIVAL_JA: list[tuple[str, str, str, str, str, str, str]] = [
    # ── 인사 (8) ──
    ("인사", "こんにちは", "こんにちは", "konnichiwa", "Hello.", "안녕하세요(낮)", "만능 낮인사·통화 시작"),
    ("인사", "おはようございます", "おはようございます", "ohayō gozaimasu", "Good morning.", "안녕하세요(아침)", "아침 인사"),
    ("인사", "こんばんは", "こんばんは", "konbanwa", "Good evening.", "안녕하세요(저녁)", "저녁 인사"),
    ("인사", "さようなら", "さようなら", "sayōnara", "Goodbye.", "안녕히 가세요", "헤어질 때"),
    ("인사", "またね", "またね", "mata ne", "See you.", "또 봐요", "가벼운 작별"),
    ("인사", "お元気ですか", "おげんきですか", "ogenki desu ka", "How are you?", "잘 지냈어요?", "안부 묻기"),
    ("인사", "いってきます", "いってきます", "ittekimasu", "I'm off.", "다녀오겠습니다", "나갈 때"),
    ("인사", "いらっしゃいませ", "いらっしゃいませ", "irasshaimase", "Welcome.", "어서 오세요", "가게에서 듣기"),
    # ── 감사·사과 (6) ──
    ("감사·사과", "ありがとうございます", "ありがとうございます", "arigatō gozaimasu", "Thank you.", "감사합니다", "격식 감사"),
    ("감사·사과", "どうも", "どうも", "dōmo", "Thanks.", "고마워요", "가벼운 감사"),
    ("감사·사과", "どういたしまして", "どういたしまして", "dō itashimashite", "You're welcome.", "천만에요", "감사 응답"),
    ("감사·사과", "すみません", "すみません", "sumimasen", "Sorry./Excuse me.", "죄송합니다·저기요", "사과·말 걸기"),
    ("감사·사과", "ごめんなさい", "ごめんなさい", "gomen nasai", "I'm sorry.", "미안해요", "가벼운 사과"),
    ("감사·사과", "大丈夫です", "だいじょうぶです", "daijōbu desu", "It's okay.", "괜찮아요", "사과 응답·사양"),
    # ── 긍정·부정·반응 (8) ──
    ("긍정·부정·반응", "はい", "はい", "hai", "Yes.", "네", "긍정"),
    ("긍정·부정·반응", "いいえ", "いいえ", "iie", "No.", "아니요", "부정"),
    ("긍정·부정·반응", "いいですね", "いいですね", "ii desu ne", "Sounds good.", "좋아요", "승낙·호감"),
    ("긍정·부정·반응", "そうです", "そうです", "sō desu", "That's right.", "맞아요", "동의"),
    ("긍정·부정·반응", "わかりました", "わかりました", "wakarimashita", "Got it.", "알겠어요", "이해 확인"),
    ("긍정·부정·반응", "わかりません", "わかりません", "wakarimasen", "I don't know.", "몰라요", "모를 때"),
    ("긍정·부정·반응", "本当ですか", "ほんとうですか", "hontō desu ka", "Really?", "진짜요?", "놀람"),
    ("긍정·부정·반응", "おいしいです", "おいしいです", "oishii desu", "It's delicious.", "맛있어요", "음식 리액션"),
    # ── 자기소개 (5) ──
    ("자기소개", "◯◯です", "◯◯です", "◯◯ desu", "I'm ◯◯.", "저는 ◯◯이에요", "이름 말하기(슬롯)"),
    ("자기소개", "◯◯から来ました", "◯◯からきました", "◯◯ kara kimashita", "I'm from ◯◯.", "◯◯에서 왔어요", "출신(슬롯)"),
    ("자기소개", "お名前は何ですか", "おなまえはなんですか", "onamae wa nan desu ka", "What's your name?", "이름이 뭐예요?", "상대 이름 묻기"),
    ("자기소개", "はじめまして", "はじめまして", "hajimemashite", "How do you do.", "처음 뵙겠습니다", "격식 첫인사"),
    ("자기소개", "よろしくお願いします", "よろしくおねがいします", "yoroshiku onegai shimasu", "Nice to meet you.", "잘 부탁드립니다", "첫 만남 마무리"),
    # ── 생존 요청 (8) ──
    ("생존 요청", "これをください", "これをください", "kore o kudasai", "This one, please.", "이거 주세요", "주문"),
    ("생존 요청", "いくらですか", "いくらですか", "ikura desu ka", "How much is it?", "얼마예요?", "가격"),
    ("생존 요청", "トイレはどこですか", "トイレはどこですか", "toire wa doko desu ka", "Where is the toilet?", "화장실이 어디예요?", "장소"),
    ("생존 요청", "助けてください", "たすけてください", "tasukete kudasai", "Please help me.", "도와주세요", "긴급"),
    ("생존 요청", "お会計お願いします", "おかいけいおねがいします", "okaikei onegai shimasu", "Check, please.", "계산해 주세요", "식당 계산"),
    ("생존 요청", "ちょっと待ってください", "ちょっとまってください", "chotto matte kudasai", "Just a moment.", "잠시만요", "시간 벌기"),
    ("생존 요청", "お腹がすきました", "おなかがすきました", "onaka ga sukimashita", "I'm hungry.", "배고파요", "상태"),
    ("생존 요청", "お水をください", "おみずをください", "omizu o kudasai", "Water, please.", "물 주세요", "식당"),
    # ── 학습자 전략 (7) ──
    ("학습자 전략", "もう一度お願いします", "もういちどおねがいします", "mō ichido onegai shimasu", "Once more, please.", "다시 말해 주세요", "못 알아들음"),
    ("학습자 전략", "ゆっくり話してください", "ゆっくりはなしてください", "yukkuri hanashite kudasai", "Please speak slowly.", "천천히 말해 주세요", "속도 조절"),
    ("학습자 전략", "よく聞こえませんでした", "よくきこえませんでした", "yoku kikoemasen deshita", "I couldn't hear well.", "잘 못 들었어요", "청취 실패"),
    ("학습자 전략", "何と言いましたか", "なんといいましたか", "nan to iimashita ka", "What did you say?", "뭐라고요?", "되묻기"),
    ("학습자 전략", "◯◯は何ですか", "◯◯はなんですか", "◯◯ wa nan desu ka", "What is ◯◯?", "◯◯이 뭐예요?", "단어 뜻(슬롯)"),
    ("학습자 전략", "日本語で何と言いますか", "にほんごでなんといいますか", "nihongo de nan to iimasu ka", "How do you say it in Japanese?", "일본어로 어떻게 말해요?", "표현 묻기"),
    ("학습자 전략", "わかりませんでした", "わかりませんでした", "wakarimasen deshita", "I didn't understand.", "이해 못 했어요", "이해 실패"),
    # ── 숫자 (4) ──
    ("숫자", "一・二・三・四・五", "いち・に・さん・し・ご", "ichi ni san shi go", "1–5", "일~오", "숫자 세기(음독)"),
    ("숫자", "六・七・八・九・十", "ろく・しち・はち・きゅう・じゅう", "roku shichi hachi kyū jū", "6–10", "육~십", "숫자 세기(음독)"),
    ("숫자", "一つ・二つ・三つ", "ひとつ・ふたつ・みっつ", "hitotsu futatsu mittsu", "one, two, three (things)", "한 개·두 개·세 개", "개수(훈독)"),
    ("숫자", "一つください", "ひとつください", "hitotsu kudasai", "One, please.", "한 개 주세요", "수량 실전"),
]


def build_survival() -> dict:
    """생존청크 46 → survival_chunks.json (한국어와 동일 스키마 + reading·meaning_ko)."""
    items = [
        {"no": i, "category": c, "surface": s, "reading": r, "roman": rm,
         "meaning_en": en, "meaning_ko": ko, "situation": sit}
        for i, (c, s, r, rm, en, ko, sit) in enumerate(_SURVIVAL_JA, start=1)
    ]
    return {"level_no": 1, "count": len(items), "language": "ja", "items": items}


def build_profiles() -> dict:
    """레벨 1(생존)~13(C4) 프로파일. profile 본문은 문법 앵커 근거 저작(PROFILE_TEXT)."""
    levels = [{
        "level_no": 1, "band": "생존", "grade": None,
        "stage_name": STAGE_NAME[1], "textbook": None,
        "profile": PROFILE_TEXT[1],
    }]
    for stage in CEFR_STAGES:
        level_no = STAGE_TO_LEVEL[stage]
        band = _band(level_no)
        levels.append({
            "level_no": level_no,
            "band": {1: "초급", 2: "중급", 3: "고급"}[band],
            "grade": {1: "A", 2: "B", 3: "C"}[band],
            "stage_name": STAGE_NAME[level_no],
            "textbook": f"JP-{stage}",
            "profile": PROFILE_TEXT[level_no],
        })
    return {
        "_comment": "일본어 레벨 프로파일(T5 저작). profile 본문은 실제 시드된 문법 앵커 근거. "
                    "level_no 축은 한국어와 동일(1=생존, 2~13=A1~C4). 생존청크는 T5 후속.",
        "language": "ja",
        "levels": levels,
    }


# ---------------------------------------------------------------- 검증·저장

def _strip_internal(items: list[dict]) -> None:
    for it in items:
        it.pop("_si", None)
        it.pop("_no", None)
        it.pop("_seq", None)


def validate(grammar: list[dict], vocab: list[dict], survival: dict) -> None:
    checks: list[tuple[str, bool]] = []

    def check(name: str, ok: bool) -> None:
        checks.append((name, ok))

    schunks = survival["items"]
    check("생존청크 46개", len(schunks) == 46 and [c["no"] for c in schunks] == list(range(1, 47)))
    check("생존청크 필드 완비(surface·reading·meaning_ko)", all(
        c.get("surface") and c.get("reading") and c.get("meaning_ko") for c in schunks
    ))
    gkeys = Counter(g["source_key"] for g in grammar)
    vkeys = Counter(v["source_key"] for v in vocab)
    check(f"문법 source_key 유일 ({len(grammar)}건)", all(c == 1 for c in gkeys.values()))
    check(f"어휘 source_key 유일 ({len(vocab)}건)", all(c == 1 for c in vkeys.values()))
    check("전 항목 level_no ∈ 2..13", all(2 <= x["level_no"] <= 13 for x in grammar + vocab))
    check("어휘 읽기(가나) 채움율 ≥ 95%",
          sum(1 for v in vocab if v.get("reading")) / max(1, len(vocab)) >= 0.95)
    check("어휘 예문 채움율 100%", all(v.get("example") for v in vocab))
    check("문법 예문 채움율 ≥ 95%",
          sum(1 for g in grammar if g.get("examples")) / max(1, len(grammar)) >= 0.95)

    failed = [n for n, ok in checks if not ok]
    for n, ok in checks:
        print(f"  {'OK ' if ok else 'FAIL'} {n}")
    if failed:
        raise AssertionError(f"검증 실패 {len(failed)}건: {failed}")


def _dump(path: Path, payload) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"생성: {path.relative_to(_ROOT)}")


def main() -> None:
    if not SENT_XLSX.exists():
        raise SystemExit(f"입력 없음: {SENT_XLSX} (level/05.다른 언어 CEFR/ 데이터 필요)")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("[1/4] 읽기·한국어뜻 조인 소스 로드")
    reading, gloss = build_reading_gloss()
    print(f"  표제어 {len(reading)} (읽기) / {len(gloss)} (한국어뜻)")

    print("[2/4] 어휘 파싱 (CEFR 문장통합본 타깃어휘)")
    vocab = parse_vocab(reading, gloss)
    vdist = Counter(v["level_no"] for v in vocab)
    print(f"  어휘 {len(vocab)}건 | 레벨분포:", {lv: vdist[lv] for lv in range(2, 14)})

    print("[3/4] 문법 파싱 (jp_grammar_12.json)")
    grammar = parse_grammar()
    gdist = Counter(g["level_no"] for g in grammar)
    print(f"  문법 {len(grammar)}건 | 레벨분포:", {lv: gdist[lv] for lv in range(2, 14)})

    print("[4/4] 저장 + 검증")
    _strip_internal(grammar)
    _strip_internal(vocab)
    grammar.sort(key=lambda g: (g["level_no"], g["seq_no"]))
    vocab.sort(key=lambda v: (v["level_no"], v["seq_no"]))
    meta = {"source": "level/05.다른 언어 CEFR/jp_cefr_rebuild",
            "generated_by": "scripts/curriculum/parse_jp.py", "language": "ja"}
    _dump(OUT_DIR / "grammar.json",
          {**meta, "assign_rule": ASSIGN_RULE, "count": len(grammar), "items": grammar})
    _dump(OUT_DIR / "vocab.json",
          {**meta, "assign_rule": ASSIGN_RULE, "count": len(vocab), "items": vocab})
    survival = build_survival()
    _dump(OUT_DIR / "survival_chunks.json", survival)
    _dump(PROFILES_JSON, build_profiles())
    validate(grammar, vocab, survival)
    print("검증 통과 — 일본어 curriculum_v2_ja 산출 완료")


if __name__ == "__main__":
    main()
