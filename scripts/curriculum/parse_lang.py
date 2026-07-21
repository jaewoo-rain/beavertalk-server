"""다른 언어(cn/en/fr/vi) 커리큘럼 raw → 정규화 JSON — parse_jp 의 일반화(멀티랭귀지 T4b).

각 언어의 XX_CEFR_문장_통합.xlsx(공통 코어 컬럼) + xx_grammar_12.json 을 읽어 한국어(cefr_v1)와
동일한 중간포맷을 산출한다. 컬럼명 편차는 LANG_CONFIG 로 흡수. (jp 는 파일 구조가 달라 parse_jp.py 별도.)

공통 원칙(한국어·일본어와 동일):
    - 어휘 = 문장통합의 타깃어휘 dedup(최저 CEFR 단계 승) → level_no=단계+1(A1→2..C4→13), example=그 문장,
      meanings.ko=한국어뜻(문장통합 내장). reading 은 이 4개 언어엔 별도 표음열이 없어 None(cn 병음은 후속).
    - 문법 = grammar_12 의 (단계, 문법구조) → dedup(최저단계 승). en 만 EGP 포맷(cefr 필드).

실행: PYTHONIOENCODING=utf-8 conda run -n beavertalk-server python scripts/curriculum/parse_lang.py <lang>
      (lang: cn|en|fr|vi|all)
산출: assets/level/curriculum_v2_<lang>/{grammar,vocab}.json + level_profiles_<lang>.json(골격)
"""

from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parent.parent.parent
_SRC = _ROOT / "level" / "05.다른 언어 CEFR"

CEFR_STAGES = ("A1", "A2", "A3", "A4", "B1", "B2", "B3", "B4", "C1", "C2", "C3", "C4")
STAGE_TO_LEVEL = {s: i + 2 for i, s in enumerate(CEFR_STAGES)}
STAGE_IDX = {s: i for i, s in enumerate(CEFR_STAGES)}

# 언어별 설정 — 컬럼명·문법 소스 편차 흡수. gloss_col: 문장통합의 한국어뜻 열 이름.
# grammar: {file, stage(단계 키), surface(구조 키), example(예문 키|None), unit, gtype}.
LANG_CONFIG: dict[str, dict] = {
    # 중국어 = ISO 639-1 'zh'(스펙·Flutter와 통일). 소스 폴더명은 cn_cefr_rebuild 유지.
    "zh": {"label": "중국어", "folder": "cn_cefr_rebuild", "gloss_col": "타깃어휘 한국어뜻",
           "grammar": {"file": "cn_grammar_12.json", "stage": "문법단계", "surface": "문법구조",
                       "example": None, "unit": "문법단원", "gtype": "분류"}},
    # 영어 문법은 별도 en_grammar_12.json(EGP·6단계·구멍)이 아니라 **문장통합**에서 뽑는다
    # — 문장통합의 문법단계(A1~C4)·문법구조가 12단계 전부를 덮는다(구멍 없음).
    "en": {"label": "영어", "folder": "en_cefr_rebuild", "gloss_col": "한국어뜻",
           "grammar_from_sentences": True},
    "fr": {"label": "프랑스어", "folder": "fr_cefr_rebuild", "gloss_col": "한국어뜻",
           "grammar": {"file": "fr_grammar_12.json", "stage": "문법단계", "surface": "문법구조",
                       "example": None, "unit": "문법단원", "gtype": "분류"}},
    "vi": {"label": "베트남어", "folder": "vi_cefr_rebuild", "gloss_col": "한국어뜻",
           "grammar": {"file": "vi_grammar_12.json", "stage": "문법단계", "surface": "문법구조",
                       "example": "예문", "unit": "문법단원", "gtype": "분류"}},
}
GRAMMAR_CORE_CAP = 45
VOCAB_CORE_CAP = 100

STAGE_NAME = {
    1: "생존 회화", 2: "초급 1 (A1)", 3: "초급 2 (A2)", 4: "초급 3 (A3)", 5: "초급 4 (A4)",
    6: "중급 1 (B1)", 7: "중급 2 (B2)", 8: "중급 3 (B3)", 9: "중급 4 (B4)",
    10: "고급 1 (C1)", 11: "고급 2 (C2)", 12: "고급 3 (C3)", 13: "고급 4 (C4)",
}

# 언어별 저작 콘텐츠(T5). 있으면 build_profiles/build_survival 이 placeholder 대신 이걸 쓴다.
# 한국어 메타 + 대상 언어 형태 인용(일본어 방식과 동일). 도그푸딩 대상=한국인(L1=ko).
PROFILE_TEXT: dict[str, dict[int, str]] = {
    "en": {
        1: "인사·숫자·자기소개·정형 표현 46개를 통째로 익히는 생존 단계. 문법 없이 "
           "'Hello', 'Thank you', 'Excuse me', 'How much is it?' 같은 덩어리 표현을 상황별로 말한다.",
        2: "가장 기초. be동사('I am...', 'This is...'), 관사(a/an/the), 대명사, 현재형"
           "('I like...', 'I have...'), 복수형까지. 주어+동사+목적어 짧은 단문으로 말한다. 과거·미래는 아직.",
        3: "과거·미래 시작. 과거형(규칙 -ed·불규칙 'went/had'), be going to 미래, can/can't, "
           "비교급('-er', 'more'), there is/are. 어제 한 일·간단한 계획을 말한다.",
        4: "진행·완료 도입. 현재진행('I'm -ing'), 현재완료(경험 'Have you ever...'), have to/should, "
           "빈도부사. 지금 하는 일·경험을 시제로 구분한다.",
        5: "미래·조건 기초. will 미래, 1형 조건문('If..., will...'), 현재완료 vs 과거, 최상급, "
           "동명사/부정사('want to', 'enjoy -ing'). 일상 대부분을 시제 구분해 말한다.",
        6: "완료진행·수동·관계절. 현재완료진행, 과거진행, 2형 조건문('If I were...'), 관계대명사"
           "(who/which/that), 수동태(현재·과거), 간접화법 도입. 문장이 길어지고 시제가 다양해진다.",
        7: "가정·완료·양태. 3형 조건문, 과거완료, 전 시제 수동태, 추측 조동사(must/might/can't be), "
           "비제한 관계절, 간접화법 완성. 복문·가정·추론을 자유롭게 구사한다.",
        8: "혼합 가정·사역. 혼합 조건문, wish/if only, 사역(have/get something done), 분사구문, "
           "고급 조동사. 복잡한 가정·뉘앙스를 표현한다.",
        9: "도치·강조 도입. 기본 도치, 분열문(It/What cleft), 미래완료/진행, 고급 수동, 가정법. "
           "강조·격식을 조절해 말한다.",
        10: "고급 도치·담화. 부정어 도치('Never have I...'), 담화 표지, 미묘한 양태, 생략, 전치(fronting). "
            "추상·전문 주제를 논리적으로 길게 전개한다.",
        11: "격식·수사. 복합 분열문, 고급 도치, 완곡·hedging, 명사화, 격식체. 뉴스·논설 수준의 정교한 영어.",
        12: "관용·정교. 관용 표현, 미묘한 강조, 문어적 수사. 원어민 서면 수준.",
        13: "최상급 — 원어민급. 수사·완곡·반어·레지스터를 자유자재로. 문학·전문 담화 수준의 완성된 영어.",
    },
    "zh": {
        1: "인사·숫자·자기소개·정형 표현 46개를 통째로 익히는 생존 단계. 문법 없이 "
           "'你好(nǐ hǎo)', '谢谢(xièxie)', '多少钱(duōshao qián)?' 같은 덩어리 표현을 상황별로 말한다.",
        2: "가장 기초. '是(shì)' 판단문·'有(yǒu)' 존재, 방위·지시·인칭대명사, 양사(量词), 숫자, "
           "'吗?' 의문. 주어+술어+목적어 짧은 단문. 시제 표지·복잡한 보어는 아직.",
        3: "완료·능원. '了(le)' 완료, 능원동사 '想/要/会/能', 비교 '比(bǐ)', '因为...所以', "
           "정도부사. 陈述句·疑问句를 구분해 말한다.",
        4: "진행·경험·중첩. 동사중첩('看看'), '过(guo)' 경험, 동량사·시량사, 방식부사, 형용사중첩. "
           "지금 하는 일·경험을 구분한다.",
        5: "보어·연결. 결과보어·추향보어(来/去), '还是...吧', '又...又...', '(在)...以前/以后'. "
           "일상 대부분을 보어로 표현한다.",
        6: "피동·이합. '被(bèi)' 피동, 离合词, 양사중첩, 접사(第-·老-·小-). 施事·受事를 명확히 한다.",
        7: "가능보어·강조. 가능보어, '越...越...', '一...也/都+不/没', 정도보어. 복문·강조를 구사한다.",
        8: "관용·복잡보어. 차용양사, '一+量词+比+一+量词', '(自)...以来', '在...方面/上/下'. 관용 표현 혼용.",
        9: "양보·강조복문. 让步复句, 반문구(反问句), 이중부정 강조, '连...也/都...' 강조. 추론·강조를 자유롭게.",
        10: "문어·격식. '从...来看', '到...为止', '拿...来说', '在...看来'. 추상 주제를 논리적으로 길게.",
        11: "관용·수사. 유사접사(超-·-化·-式), '为了...而...', '非...不可' 강조. 논설 수준의 정교한 중국어.",
        12: "고급 관용. '所谓...就是...', '无非...而已', '以...为...', '因...而...'. 원어민 서면 수준.",
        13: "최상급 — 원어민급. '话又说回来', 'X了又Y', '别提多X了' 등 구어·문어 수사 자유자재. 완성된 중국어.",
    },
    "fr": {
        1: "인사·숫자·자기소개·정형 표현 46개를 통째로 익히는 생존 단계. 문법 없이 "
           "'Bonjour', 'Merci', 'Combien ça coûte?' 같은 덩어리 표현을 상황별로 말한다.",
        2: "가장 기초. être/avoir ('Je suis...', 'J'ai...'), 관사(le/la/un/une), -er 동사 현재형, "
           "형용사 성수일치, 소유형용사(mon/ma/mes). 주어+동사+보어 짧은 단문. 과거·복잡한 시제는 아직.",
        3: "과거·근접미래. passé composé ('J'ai mangé'), futur proche (aller+inf), 부정 (ne...pas), "
           "비교 (plus...que). 어제 한 일·계획을 말한다.",
        4: "반과거·대명사. imparfait, 목적격 대명사(COD/COI: le/la/lui), 재귀동사(se laver). "
           "과거를 묘사·구분한다.",
        5: "단순미래·조건 기초. futur simple, 1형 가정 (si+présent), 관계대명사(qui/que), "
           "부분관사(du/de la). 일상 대부분을 시제로 표현한다.",
        6: "접속법·조건법 도입. subjonctif présent(il faut que), conditionnel présent, "
           "관계대명사(dont/où), passé composé vs imparfait 구분. 문장이 길어진다.",
        7: "접속법·가정 완성. subjonctif 완전, plus-que-parfait, 수동태, 2형 가정 (si+imparfait), "
           "gérondif(en -ant). 복문·가정을 자유롭게 구사한다.",
        8: "조건 과거·복잡 시제. conditionnel passé, subjonctif passé, 3형 가정, 복잡 관계절. "
           "뉘앙스·후회를 표현한다.",
        9: "격식·강조. mise en relief (c'est...que), 도치 의문, 고급 수동, 담화 연결. 격식·강조를 조절.",
        10: "문어·논리. connecteurs logiques, subjonctif 뉘앙스, participe présent. 추상 담화를 길게.",
        11: "격식·수사. 명사화, 완곡, 격식체, 복잡 종속. 논설 수준의 정교한 프랑스어.",
        12: "고급 관용·문어. passé simple(문어 시제), 관용 표현, 미묘한 강조. 원어민 서면 수준.",
        13: "최상급 — 원어민급. 수사·완곡·반어·레지스터 자유자재. 문학·전문 담화 수준의 완성된 프랑스어.",
    },
    "vi": {
        1: "인사·숫자·자기소개·정형 표현 46개를 통째로 익히는 생존 단계. 문법 없이 "
           "'Xin chào', 'Cảm ơn', 'Bao nhiêu tiền?' 같은 덩어리 표현을 상황별로 말한다.",
        2: "가장 기초. 'là(be)' 판단문·'có' 존재, 인칭대명사, 분류사(classifier), 숫자, "
           "'có...không?' 의문, 'cũng/đều'. 주어+동사+목적어 짧은 단문. 시제 표지는 아직.",
        3: "완료·시간. 'đã...chưa?' 완료, 시간 의문(thứ mấy/bao lâu), 'à/chứ' 의문. "
           "어제 한 일을 시간과 함께 말한다.",
        4: "의도·비교. 'muốn/định' 원함·의도, 'bằng' 비교, 복수(những/các), 'ai cũng'(누구나). "
           "일상 요청·비교를 말한다.",
        5: "미래·이유. 'sắp'(막), 'vừa...vừa', 'vì...nên'(때문에), 'hình như...thì phải'(추측). "
           "계획·이유를 표현한다.",
        6: "피동·강조. 'được'(가능/피동), 'tự...lấy'(스스로), 어기조사(nhỉ/nhé), 'hóa ra'(알고 보니). "
           "施事·강조를 명확히 한다.",
        7: "중첩·관용. 중첩(sáng sáng), 'nói chung/riêng', 'một mặt...mặt khác'. 관용 표현 혼용.",
        8: "명사화·복잡 의문. 'trừ/kể cả', 'sự+동사' 명사화, 복잡 의문(Sao mà...thế). 추상 전개.",
        9: "강조·양보. 'quả là'(정말), 'huống chi'(하물며), 중첩 강조(nhà nhà). 강조·추론을 자유롭게.",
        10: "문어·격식. 'kẻ...người...', 부정확 수량 표현, 격식 구조. 추상 주제를 길게.",
        11: "관용·수사. 'biết đâu đấy', 'liệu+', 담화 표지(nói tóm lại). 논설 수준.",
        12: "고급 관용. 복잡 구문(A còn...nữa là B), 동사 변별(mời/nhờ/khuyên). 원어민 서면 수준.",
        13: "최상급 — 원어민급. 'Cứ+동사+đi', 'dù sao...cũng' 등 구어·문어 수사 자유자재. 완성된 베트남어.",
    },
}

# (category, surface[대상어], reading, roman, meaning_en, meaning_ko, situation[한국어])
SURVIVAL: dict[str, list[tuple]] = {
    "en": [
        ("인사", "Hello.", None, None, "Hello.", "안녕하세요", "만능 인사·통화 시작"),
        ("인사", "Hi, nice to meet you.", None, None, "Nice to meet you.", "안녕, 만나서 반가워요", "첫 만남"),
        ("인사", "How are you?", None, None, "How are you?", "잘 지내요?", "안부 묻기"),
        ("인사", "Good morning.", None, None, "Good morning.", "좋은 아침이에요", "아침 인사"),
        ("인사", "Goodbye.", None, None, "Goodbye.", "안녕히 가세요", "헤어질 때"),
        ("인사", "See you later!", None, None, "See you later.", "또 봐요", "가벼운 작별"),
        ("인사", "Have a nice day.", None, None, "Have a nice day.", "좋은 하루 보내세요", "작별 덧붙임"),
        ("인사", "Welcome!", None, None, "Welcome.", "어서 오세요", "맞이할 때"),
        ("감사·사과", "Thank you.", None, None, "Thank you.", "감사합니다", "감사"),
        ("감사·사과", "Thanks a lot.", None, None, "Thanks a lot.", "정말 고마워요", "강한 감사"),
        ("감사·사과", "You're welcome.", None, None, "You're welcome.", "천만에요", "감사 응답"),
        ("감사·사과", "I'm sorry.", None, None, "I'm sorry.", "죄송합니다", "사과"),
        ("감사·사과", "Excuse me.", None, None, "Excuse me.", "실례합니다", "말 걸기·사과"),
        ("감사·사과", "That's okay.", None, None, "That's okay.", "괜찮아요", "사과 응답·사양"),
        ("긍정·부정·반응", "Yes.", None, None, "Yes.", "네", "긍정"),
        ("긍정·부정·반응", "No.", None, None, "No.", "아니요", "부정"),
        ("긍정·부정·반응", "Sounds good.", None, None, "Sounds good.", "좋아요", "승낙"),
        ("긍정·부정·반응", "That's right.", None, None, "That's right.", "맞아요", "동의"),
        ("긍정·부정·반응", "I got it.", None, None, "I got it.", "알겠어요", "이해 확인"),
        ("긍정·부정·반응", "I don't know.", None, None, "I don't know.", "몰라요", "모를 때"),
        ("긍정·부정·반응", "Really?", None, None, "Really?", "진짜요?", "놀람"),
        ("긍정·부정·반응", "It's delicious.", None, None, "It's delicious.", "맛있어요", "음식 리액션"),
        ("자기소개", "I'm ◯◯.", None, None, "I'm ◯◯.", "저는 ◯◯이에요", "이름(슬롯)"),
        ("자기소개", "I'm from ◯◯.", None, None, "I'm from ◯◯.", "◯◯에서 왔어요", "출신(슬롯)"),
        ("자기소개", "What's your name?", None, None, "What's your name?", "이름이 뭐예요?", "상대 이름 묻기"),
        ("자기소개", "How do you do.", None, None, "How do you do.", "처음 뵙겠습니다", "격식 첫인사"),
        ("자기소개", "Nice to meet you.", None, None, "Nice to meet you.", "잘 부탁해요", "첫 만남 마무리"),
        ("생존 요청", "This one, please.", None, None, "This one, please.", "이거 주세요", "주문"),
        ("생존 요청", "How much is it?", None, None, "How much is it?", "얼마예요?", "가격"),
        ("생존 요청", "Where is the bathroom?", None, None, "Where is the bathroom?", "화장실이 어디예요?", "장소"),
        ("생존 요청", "Please help me.", None, None, "Please help me.", "도와주세요", "긴급"),
        ("생존 요청", "Excuse me! (calling)", None, None, "Excuse me!", "여기요!", "점원 부르기"),
        ("생존 요청", "Just a moment.", None, None, "Just a moment.", "잠시만요", "시간 벌기"),
        ("생존 요청", "I'm hungry.", None, None, "I'm hungry.", "배고파요", "상태"),
        ("생존 요청", "Water, please.", None, None, "Water, please.", "물 주세요", "식당"),
        ("학습자 전략", "Could you say that again?", None, None, "Say that again, please.", "다시 말해 주세요", "못 알아들음"),
        ("학습자 전략", "Please speak slowly.", None, None, "Please speak slowly.", "천천히 말해 주세요", "속도 조절"),
        ("학습자 전략", "I didn't catch that.", None, None, "I didn't catch that.", "잘 못 들었어요", "청취 실패"),
        ("학습자 전략", "What did you say?", None, None, "What did you say?", "뭐라고요?", "되묻기"),
        ("학습자 전략", "What does ◯◯ mean?", None, None, "What does ◯◯ mean?", "◯◯이 무슨 뜻이에요?", "단어 뜻(슬롯)"),
        ("학습자 전략", "How do you say it in English?", None, None, "How do you say it in English?", "영어로 어떻게 말해요?", "표현 묻기"),
        ("학습자 전략", "I don't understand.", None, None, "I don't understand.", "이해 못 했어요", "이해 실패"),
        ("숫자", "one, two, three, four, five", None, None, "1–5", "일~오", "숫자 세기"),
        ("숫자", "six, seven, eight, nine, ten", None, None, "6–10", "육~십", "숫자 세기"),
        ("숫자", "first, second, third", None, None, "1st, 2nd, 3rd", "첫째·둘째·셋째", "서수"),
        ("숫자", "One, please.", None, None, "One, please.", "하나 주세요", "수량 실전"),
    ],
    "zh": [
        ("인사", "你好。", "nǐ hǎo", "nǐ hǎo", "Hello.", "안녕하세요", "만능 인사·통화 시작"),
        ("인사", "很高兴认识你。", "hěn gāoxìng rènshi nǐ", "hěn gāoxìng rènshi nǐ", "Nice to meet you.", "만나서 반가워요", "첫 만남"),
        ("인사", "你好吗？", "nǐ hǎo ma", "nǐ hǎo ma", "How are you?", "잘 지내요?", "안부 묻기"),
        ("인사", "早上好。", "zǎoshang hǎo", "zǎoshang hǎo", "Good morning.", "좋은 아침이에요", "아침 인사"),
        ("인사", "再见。", "zàijiàn", "zàijiàn", "Goodbye.", "안녕히 가세요", "헤어질 때"),
        ("인사", "回头见！", "huítóu jiàn", "huítóu jiàn", "See you later.", "또 봐요", "가벼운 작별"),
        ("인사", "祝你今天愉快。", "zhù nǐ jīntiān yúkuài", "zhù nǐ jīntiān yúkuài", "Have a nice day.", "좋은 하루 보내세요", "작별 덧붙임"),
        ("인사", "欢迎光临！", "huānyíng guānglín", "huānyíng guānglín", "Welcome.", "어서 오세요", "가게에서 듣기"),
        ("감사·사과", "谢谢。", "xièxie", "xièxie", "Thank you.", "감사합니다", "감사"),
        ("감사·사과", "非常感谢。", "fēicháng gǎnxiè", "fēicháng gǎnxiè", "Thanks a lot.", "정말 고마워요", "강한 감사"),
        ("감사·사과", "不客气。", "bú kèqi", "bú kèqi", "You're welcome.", "천만에요", "감사 응답"),
        ("감사·사과", "对不起。", "duìbuqǐ", "duìbuqǐ", "I'm sorry.", "죄송합니다", "사과"),
        ("감사·사과", "不好意思。", "bù hǎoyìsi", "bù hǎoyìsi", "Excuse me.", "실례합니다", "말 걸기·사과"),
        ("감사·사과", "没关系。", "méi guānxi", "méi guānxi", "That's okay.", "괜찮아요", "사과 응답"),
        ("긍정·부정·반응", "是的。", "shì de", "shì de", "Yes.", "네", "긍정"),
        ("긍정·부정·반응", "不是。", "bú shì", "bú shì", "No.", "아니요", "부정"),
        ("긍정·부정·반응", "好的。", "hǎo de", "hǎo de", "Sounds good.", "좋아요", "승낙"),
        ("긍정·부정·반응", "对。", "duì", "duì", "That's right.", "맞아요", "동의"),
        ("긍정·부정·반응", "我知道了。", "wǒ zhīdào le", "wǒ zhīdào le", "I got it.", "알겠어요", "이해 확인"),
        ("긍정·부정·반응", "我不知道。", "wǒ bù zhīdào", "wǒ bù zhīdào", "I don't know.", "몰라요", "모를 때"),
        ("긍정·부정·반응", "真的吗？", "zhēn de ma", "zhēn de ma", "Really?", "진짜요?", "놀람"),
        ("긍정·부정·반응", "很好吃。", "hěn hǎochī", "hěn hǎochī", "It's delicious.", "맛있어요", "음식 리액션"),
        ("자기소개", "我叫◯◯。", "wǒ jiào ◯◯", "wǒ jiào ◯◯", "I'm ◯◯.", "저는 ◯◯이에요", "이름(슬롯)"),
        ("자기소개", "我是◯◯人。", "wǒ shì ◯◯ rén", "wǒ shì ◯◯ rén", "I'm from ◯◯.", "◯◯에서 왔어요", "출신(슬롯)"),
        ("자기소개", "你叫什么名字？", "nǐ jiào shénme míngzi", "nǐ jiào shénme míngzi", "What's your name?", "이름이 뭐예요?", "상대 이름 묻기"),
        ("자기소개", "初次见面。", "chūcì jiànmiàn", "chūcì jiànmiàn", "How do you do.", "처음 뵙겠습니다", "격식 첫인사"),
        ("자기소개", "请多关照。", "qǐng duō guānzhào", "qǐng duō guānzhào", "Nice to meet you.", "잘 부탁해요", "첫 만남 마무리"),
        ("생존 요청", "请给我这个。", "qǐng gěi wǒ zhège", "qǐng gěi wǒ zhège", "This one, please.", "이거 주세요", "주문"),
        ("생존 요청", "多少钱？", "duōshao qián", "duōshao qián", "How much is it?", "얼마예요?", "가격"),
        ("생존 요청", "洗手间在哪儿？", "xǐshǒujiān zài nǎr", "xǐshǒujiān zài nǎr", "Where is the bathroom?", "화장실이 어디예요?", "장소"),
        ("생존 요청", "请帮帮我。", "qǐng bāngbang wǒ", "qǐng bāngbang wǒ", "Please help me.", "도와주세요", "긴급"),
        ("생존 요청", "服务员！", "fúwùyuán", "fúwùyuán", "Excuse me! (calling)", "여기요!", "점원 부르기"),
        ("생존 요청", "请稍等。", "qǐng shāoděng", "qǐng shāoděng", "Just a moment.", "잠시만요", "시간 벌기"),
        ("생존 요청", "我饿了。", "wǒ è le", "wǒ è le", "I'm hungry.", "배고파요", "상태"),
        ("생존 요청", "请给我水。", "qǐng gěi wǒ shuǐ", "qǐng gěi wǒ shuǐ", "Water, please.", "물 주세요", "식당"),
        ("학습자 전략", "请再说一遍。", "qǐng zài shuō yí biàn", "qǐng zài shuō yí biàn", "Say that again, please.", "다시 말해 주세요", "못 알아들음"),
        ("학습자 전략", "请说慢一点。", "qǐng shuō màn yìdiǎn", "qǐng shuō màn yìdiǎn", "Please speak slowly.", "천천히 말해 주세요", "속도 조절"),
        ("학습자 전략", "我没听清楚。", "wǒ méi tīng qīngchu", "wǒ méi tīng qīngchu", "I didn't catch that.", "잘 못 들었어요", "청취 실패"),
        ("학습자 전략", "你说什么？", "nǐ shuō shénme", "nǐ shuō shénme", "What did you say?", "뭐라고요?", "되묻기"),
        ("학습자 전략", "◯◯是什么意思？", "◯◯ shì shénme yìsi", "◯◯ shì shénme yìsi", "What does ◯◯ mean?", "◯◯이 무슨 뜻이에요?", "단어 뜻(슬롯)"),
        ("학습자 전략", "用中文怎么说？", "yòng zhōngwén zěnme shuō", "yòng zhōngwén zěnme shuō", "How do you say it in Chinese?", "중국어로 어떻게 말해요?", "표현 묻기"),
        ("학습자 전략", "我不明白。", "wǒ bù míngbai", "wǒ bù míngbai", "I don't understand.", "이해 못 했어요", "이해 실패"),
        ("숫자", "一、二、三、四、五", "yī èr sān sì wǔ", "yī èr sān sì wǔ", "1–5", "일~오", "숫자 세기"),
        ("숫자", "六、七、八、九、十", "liù qī bā jiǔ shí", "liù qī bā jiǔ shí", "6–10", "육~십", "숫자 세기"),
        ("숫자", "一百、一千", "yìbǎi, yìqiān", "yìbǎi, yìqiān", "100, 1000", "백·천", "큰 수"),
        ("숫자", "请给我一个。", "qǐng gěi wǒ yí ge", "qǐng gěi wǒ yí ge", "One, please.", "하나 주세요", "수량 실전"),
    ],
    "fr": [
        ("인사", "Bonjour.", None, None, "Hello.", "안녕하세요", "만능 인사·통화 시작"),
        ("인사", "Enchanté(e).", None, None, "Nice to meet you.", "만나서 반가워요", "첫 만남"),
        ("인사", "Comment allez-vous ?", None, None, "How are you?", "잘 지내요?", "안부 묻기"),
        ("인사", "Bonjour (le matin).", None, None, "Good morning.", "좋은 아침이에요", "아침 인사"),
        ("인사", "Au revoir.", None, None, "Goodbye.", "안녕히 가세요", "헤어질 때"),
        ("인사", "À bientôt !", None, None, "See you later.", "또 봐요", "가벼운 작별"),
        ("인사", "Bonne journée.", None, None, "Have a nice day.", "좋은 하루 보내세요", "작별 덧붙임"),
        ("인사", "Bienvenue !", None, None, "Welcome.", "어서 오세요", "맞이할 때"),
        ("감사·사과", "Merci.", None, None, "Thank you.", "감사합니다", "감사"),
        ("감사·사과", "Merci beaucoup.", None, None, "Thanks a lot.", "정말 고마워요", "강한 감사"),
        ("감사·사과", "De rien.", None, None, "You're welcome.", "천만에요", "감사 응답"),
        ("감사·사과", "Désolé(e).", None, None, "I'm sorry.", "죄송합니다", "사과"),
        ("감사·사과", "Excusez-moi.", None, None, "Excuse me.", "실례합니다", "말 걸기·사과"),
        ("감사·사과", "Ce n'est pas grave.", None, None, "That's okay.", "괜찮아요", "사과 응답"),
        ("긍정·부정·반응", "Oui.", None, None, "Yes.", "네", "긍정"),
        ("긍정·부정·반응", "Non.", None, None, "No.", "아니요", "부정"),
        ("긍정·부정·반응", "D'accord.", None, None, "Sounds good.", "좋아요", "승낙"),
        ("긍정·부정·반응", "C'est exact.", None, None, "That's right.", "맞아요", "동의"),
        ("긍정·부정·반응", "J'ai compris.", None, None, "I got it.", "알겠어요", "이해 확인"),
        ("긍정·부정·반응", "Je ne sais pas.", None, None, "I don't know.", "몰라요", "모를 때"),
        ("긍정·부정·반응", "Vraiment ?", None, None, "Really?", "진짜요?", "놀람"),
        ("긍정·부정·반응", "C'est délicieux.", None, None, "It's delicious.", "맛있어요", "음식 리액션"),
        ("자기소개", "Je m'appelle ◯◯.", None, None, "I'm ◯◯.", "저는 ◯◯이에요", "이름(슬롯)"),
        ("자기소개", "Je viens de ◯◯.", None, None, "I'm from ◯◯.", "◯◯에서 왔어요", "출신(슬롯)"),
        ("자기소개", "Comment vous appelez-vous ?", None, None, "What's your name?", "이름이 뭐예요?", "상대 이름 묻기"),
        ("자기소개", "Enchanté(e) de faire votre connaissance.", None, None, "How do you do.", "처음 뵙겠습니다", "격식 첫인사"),
        ("자기소개", "Ravi(e) de vous rencontrer.", None, None, "Nice to meet you.", "잘 부탁해요", "첫 만남 마무리"),
        ("생존 요청", "Celui-ci, s'il vous plaît.", None, None, "This one, please.", "이거 주세요", "주문"),
        ("생존 요청", "Combien ça coûte ?", None, None, "How much is it?", "얼마예요?", "가격"),
        ("생존 요청", "Où sont les toilettes ?", None, None, "Where is the bathroom?", "화장실이 어디예요?", "장소"),
        ("생존 요청", "Aidez-moi, s'il vous plaît.", None, None, "Please help me.", "도와주세요", "긴급"),
        ("생존 요청", "S'il vous plaît ! (appeler)", None, None, "Excuse me!", "여기요!", "점원 부르기"),
        ("생존 요청", "Un instant.", None, None, "Just a moment.", "잠시만요", "시간 벌기"),
        ("생존 요청", "J'ai faim.", None, None, "I'm hungry.", "배고파요", "상태"),
        ("생존 요청", "De l'eau, s'il vous plaît.", None, None, "Water, please.", "물 주세요", "식당"),
        ("학습자 전략", "Pouvez-vous répéter ?", None, None, "Say that again, please.", "다시 말해 주세요", "못 알아들음"),
        ("학습자 전략", "Parlez lentement, s'il vous plaît.", None, None, "Please speak slowly.", "천천히 말해 주세요", "속도 조절"),
        ("학습자 전략", "Je n'ai pas bien entendu.", None, None, "I didn't catch that.", "잘 못 들었어요", "청취 실패"),
        ("학습자 전략", "Qu'avez-vous dit ?", None, None, "What did you say?", "뭐라고요?", "되묻기"),
        ("학습자 전략", "Que veut dire ◯◯ ?", None, None, "What does ◯◯ mean?", "◯◯이 무슨 뜻이에요?", "단어 뜻(슬롯)"),
        ("학습자 전략", "Comment dit-on en français ?", None, None, "How do you say it in French?", "프랑스어로 어떻게 말해요?", "표현 묻기"),
        ("학습자 전략", "Je ne comprends pas.", None, None, "I don't understand.", "이해 못 했어요", "이해 실패"),
        ("숫자", "un, deux, trois, quatre, cinq", None, None, "1–5", "일~오", "숫자 세기"),
        ("숫자", "six, sept, huit, neuf, dix", None, None, "6–10", "육~십", "숫자 세기"),
        ("숫자", "cent, mille", None, None, "100, 1000", "백·천", "큰 수"),
        ("숫자", "Un, s'il vous plaît.", None, None, "One, please.", "하나 주세요", "수량 실전"),
    ],
    "vi": [
        ("인사", "Xin chào.", None, None, "Hello.", "안녕하세요", "만능 인사·통화 시작"),
        ("인사", "Rất vui được gặp bạn.", None, None, "Nice to meet you.", "만나서 반가워요", "첫 만남"),
        ("인사", "Bạn khỏe không?", None, None, "How are you?", "잘 지내요?", "안부 묻기"),
        ("인사", "Chào buổi sáng.", None, None, "Good morning.", "좋은 아침이에요", "아침 인사"),
        ("인사", "Tạm biệt.", None, None, "Goodbye.", "안녕히 가세요", "헤어질 때"),
        ("인사", "Hẹn gặp lại!", None, None, "See you later.", "또 봐요", "가벼운 작별"),
        ("인사", "Chúc một ngày tốt lành.", None, None, "Have a nice day.", "좋은 하루 보내세요", "작별 덧붙임"),
        ("인사", "Chào mừng!", None, None, "Welcome.", "어서 오세요", "맞이할 때"),
        ("감사·사과", "Cảm ơn.", None, None, "Thank you.", "감사합니다", "감사"),
        ("감사·사과", "Cảm ơn rất nhiều.", None, None, "Thanks a lot.", "정말 고마워요", "강한 감사"),
        ("감사·사과", "Không có gì.", None, None, "You're welcome.", "천만에요", "감사 응답"),
        ("감사·사과", "Xin lỗi.", None, None, "I'm sorry.", "죄송합니다", "사과"),
        ("감사·사과", "Xin lỗi (làm phiền).", None, None, "Excuse me.", "실례합니다", "말 걸기·사과"),
        ("감사·사과", "Không sao.", None, None, "That's okay.", "괜찮아요", "사과 응답"),
        ("긍정·부정·반응", "Vâng.", None, None, "Yes.", "네", "긍정"),
        ("긍정·부정·반응", "Không.", None, None, "No.", "아니요", "부정"),
        ("긍정·부정·반응", "Được.", None, None, "Sounds good.", "좋아요", "승낙"),
        ("긍정·부정·반응", "Đúng rồi.", None, None, "That's right.", "맞아요", "동의"),
        ("긍정·부정·반응", "Tôi hiểu rồi.", None, None, "I got it.", "알겠어요", "이해 확인"),
        ("긍정·부정·반응", "Tôi không biết.", None, None, "I don't know.", "몰라요", "모를 때"),
        ("긍정·부정·반응", "Thật à?", None, None, "Really?", "진짜요?", "놀람"),
        ("긍정·부정·반응", "Ngon quá.", None, None, "It's delicious.", "맛있어요", "음식 리액션"),
        ("자기소개", "Tôi tên là ◯◯.", None, None, "I'm ◯◯.", "저는 ◯◯이에요", "이름(슬롯)"),
        ("자기소개", "Tôi đến từ ◯◯.", None, None, "I'm from ◯◯.", "◯◯에서 왔어요", "출신(슬롯)"),
        ("자기소개", "Bạn tên là gì?", None, None, "What's your name?", "이름이 뭐예요?", "상대 이름 묻기"),
        ("자기소개", "Rất hân hạnh.", None, None, "How do you do.", "처음 뵙겠습니다", "격식 첫인사"),
        ("자기소개", "Rất vui được làm quen.", None, None, "Nice to meet you.", "잘 부탁해요", "첫 만남 마무리"),
        ("생존 요청", "Cho tôi cái này.", None, None, "This one, please.", "이거 주세요", "주문"),
        ("생존 요청", "Bao nhiêu tiền?", None, None, "How much is it?", "얼마예요?", "가격"),
        ("생존 요청", "Nhà vệ sinh ở đâu?", None, None, "Where is the bathroom?", "화장실이 어디예요?", "장소"),
        ("생존 요청", "Làm ơn giúp tôi.", None, None, "Please help me.", "도와주세요", "긴급"),
        ("생존 요청", "Em ơi!", None, None, "Excuse me! (calling)", "여기요!", "점원 부르기"),
        ("생존 요청", "Chờ một chút.", None, None, "Just a moment.", "잠시만요", "시간 벌기"),
        ("생존 요청", "Tôi đói.", None, None, "I'm hungry.", "배고파요", "상태"),
        ("생존 요청", "Cho tôi nước.", None, None, "Water, please.", "물 주세요", "식당"),
        ("학습자 전략", "Bạn nói lại được không?", None, None, "Say that again, please.", "다시 말해 주세요", "못 알아들음"),
        ("학습자 전략", "Xin nói chậm hơn.", None, None, "Please speak slowly.", "천천히 말해 주세요", "속도 조절"),
        ("학습자 전략", "Tôi nghe không rõ.", None, None, "I didn't catch that.", "잘 못 들었어요", "청취 실패"),
        ("학습자 전략", "Bạn nói gì?", None, None, "What did you say?", "뭐라고요?", "되묻기"),
        ("학습자 전략", "◯◯ nghĩa là gì?", None, None, "What does ◯◯ mean?", "◯◯이 무슨 뜻이에요?", "단어 뜻(슬롯)"),
        ("학습자 전략", "Nói tiếng Việt thế nào?", None, None, "How do you say it in Vietnamese?", "베트남어로 어떻게 말해요?", "표현 묻기"),
        ("학습자 전략", "Tôi không hiểu.", None, None, "I don't understand.", "이해 못 했어요", "이해 실패"),
        ("숫자", "một, hai, ba, bốn, năm", None, None, "1–5", "일~오", "숫자 세기"),
        ("숫자", "sáu, bảy, tám, chín, mười", None, None, "6–10", "육~십", "숫자 세기"),
        ("숫자", "trăm, nghìn", None, None, "100, 1000", "백·천", "큰 수"),
        ("숫자", "Cho tôi một cái.", None, None, "One, please.", "하나 주세요", "수량 실전"),
    ],
}


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _band(level_no: int) -> int:
    return 1 if level_no <= 5 else (2 if level_no <= 9 else 3)


def _header_map(ws) -> dict[str, int]:
    """헤더행 → {컬럼명: 0-based 인덱스}."""
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(c).strip(): i for i, c in enumerate(hdr) if c is not None}


def parse_vocab(lang: str, cfg: dict) -> list[dict]:
    """문장통합 타깃어휘 → 어휘 항목(최저 CEFR 단계 승, 예문·한국어뜻 인라인)."""
    folder = _SRC / cfg["folder"]
    xlsx = next(folder.glob("*_문장_통합.xlsx"))
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hm = _header_map(ws)
    ci = lambda name: hm.get(name)  # noqa: E731
    c_stage, c_struct, c_sent = ci("문법단계"), ci("문법구조"), ci("문장")
    c_target, c_pos, c_gloss = ci("타깃어휘"), ci("타깃품사"), ci(cfg["gloss_col"])

    chosen: dict[str, dict] = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        stage = _clean(row[c_stage]) if c_stage is not None else None
        surface = _clean(row[c_target]) if c_target is not None else None
        if not stage or not surface or stage not in STAGE_TO_LEVEL:
            continue
        si = STAGE_IDX[stage]
        prev = chosen.get(surface)
        if prev is not None and (si, i) >= (prev["_si"], prev["_no"]):
            continue
        chosen[surface] = {
            "surface": surface, "stage": stage, "_si": si, "_no": i,
            "sentence": _clean(row[c_sent]) if c_sent is not None else None,
            "pos": _clean(row[c_pos]) if c_pos is not None else None,
            "gloss": _clean(row[c_gloss]) if c_gloss is not None else None,
        }
    wb.close()

    items = []
    for c in chosen.values():
        level_no = STAGE_TO_LEVEL[c["stage"]]
        meanings = {"ko": c["gloss"]} if c["gloss"] else None
        items.append({
            # topik_grade: 언어별 어휘등급 대신 레벨에서 유도한 등급밴드(1~6) — vocab NOT NULL 제약 충족.
            "kind": "vocab", "source_key": f"v:{lang}:{c['surface']}",
            "band": _band(level_no), "topik_grade": level_no // 2, "level_no": level_no,
            "assign_rule": f"cefr_{lang}_v1", "surface": c["surface"], "reading": None,
            "pos_primary": c["pos"], "pos_list": [c["pos"]] if c["pos"] else None, "pos_raw": c["pos"],
            "is_verb_priority": False, "is_core": False,
            "example": c["sentence"], "meanings": meanings,
            "_si": c["_si"], "_no": c["_no"],
        })
    _rank_and_cap(items, VOCAB_CORE_CAP)
    return items


def parse_grammar_from_sentences(lang: str, cfg: dict) -> list[dict]:
    """문장통합의 문법단계·문법구조 → 문법 항목(surface dedup, 최저단계 승, 예문=그 문장).

    영어처럼 별도 grammar_12(EGP·구멍)를 못 쓸 때 — 문장통합이 A1~C4 12단계를 다 덮는다.
    """
    folder = _SRC / cfg["folder"]
    xlsx = next(folder.glob("*_문장_통합.xlsx"))
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hm = _header_map(ws)
    c_stage, c_struct, c_sent = hm.get("문법단계"), hm.get("문법구조"), hm.get("문장")
    c_unit, c_gtype = hm.get("문법단원"), hm.get("문법분류")
    chosen: dict[str, dict] = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        stage = _clean(row[c_stage]) if c_stage is not None else None
        surface = _clean(row[c_struct]) if c_struct is not None else None
        if not stage or not surface or stage not in STAGE_TO_LEVEL:
            continue
        si = STAGE_IDX[stage]
        prev = chosen.get(surface)
        if prev is not None and (si, i) >= (prev["_si"], prev["_seq"]):
            continue
        chosen[surface] = {
            "surface": surface, "stage": stage, "_si": si, "_seq": i,
            "unit": _clean(row[c_unit]) if c_unit is not None else None,
            "unit_title": None,
            "gtype": _clean(row[c_gtype]) if c_gtype is not None else None,
            "example": _clean(row[c_sent]) if c_sent is not None else None,
        }
    wb.close()
    items = []
    for c in chosen.values():
        level_no = STAGE_TO_LEVEL[c["stage"]]
        items.append({
            "kind": "grammar", "source_key": f"g:{lang}:{c['surface']}",
            "band": _band(level_no), "textbook_code": f"{lang.upper()}-{c['stage']}",
            "level_no": level_no, "assign_rule": f"cefr_{lang}_v1", "surface": c["surface"],
            "reading": None, "unit": c["unit"], "unit_title": c["unit_title"],
            "grammar_type": c["gtype"], "examples": [c["example"]] if c["example"] else None,
            "explanation": None, "caution": None, "is_core": False,
            "_si": c["_si"], "_seq": c["_seq"],
        })
    _rank_and_cap(items, GRAMMAR_CORE_CAP)
    return items


def parse_grammar(lang: str, cfg: dict) -> list[dict]:
    """grammar_12 → 문법 항목(surface dedup, 최저단계 승). 문장통합 소스면 그쪽 파서로 위임."""
    if cfg.get("grammar_from_sentences"):
        return parse_grammar_from_sentences(lang, cfg)
    gc = cfg["grammar"]
    data = json.loads((_SRC / cfg["folder"] / gc["file"]).read_text(encoding="utf-8"))
    chosen: dict[str, dict] = {}
    for i, e in enumerate(data):
        stage = _clean(e.get(gc["stage"]))
        surface = _clean(e.get(gc["surface"]))
        if not stage or not surface or stage not in STAGE_TO_LEVEL:
            continue
        si = STAGE_IDX[stage]
        prev = chosen.get(surface)
        if prev is not None and (si, i) >= (prev["_si"], prev["_seq"]):
            continue
        chosen[surface] = {
            "surface": surface, "stage": stage, "_si": si, "_seq": i,
            "unit": _clean(e.get(gc["unit"])) if gc.get("unit") else None,
            "unit_title": _clean(e.get("문법제목") or e.get("guideword")),
            "gtype": _clean(e.get(gc["gtype"])) if gc.get("gtype") else None,
            "example": _clean(e.get(gc["example"])) if gc.get("example") else None,
        }
    items = []
    for c in chosen.values():
        level_no = STAGE_TO_LEVEL[c["stage"]]
        items.append({
            "kind": "grammar", "source_key": f"g:{lang}:{c['surface']}",
            "band": _band(level_no), "textbook_code": f"{lang.upper()}-{c['stage']}",
            "level_no": level_no, "assign_rule": f"cefr_{lang}_v1", "surface": c["surface"],
            "reading": None, "unit": c["unit"], "unit_title": c["unit_title"],
            "grammar_type": c["gtype"], "examples": [c["example"]] if c["example"] else None,
            "explanation": None, "caution": None, "is_core": False,
            "_si": c["_si"], "_seq": c["_seq"],
        })
    _rank_and_cap(items, GRAMMAR_CORE_CAP)
    return items


def _rank_and_cap(items: list[dict], cap: int) -> None:
    from collections import defaultdict
    by_level: dict[int, list[dict]] = defaultdict(list)
    for it in items:
        by_level[it["level_no"]].append(it)
    for group in by_level.values():
        group.sort(key=lambda x: (x["_si"], x.get("_no", x.get("_seq", 0))))
        for rank, it in enumerate(group, start=1):
            it["priority_rank"] = rank
            it["seq_no"] = rank
            it["is_core"] = rank <= cap


def build_profiles(lang: str, cfg: dict) -> dict:
    """레벨 1(생존)~13(C4) 프로파일. 저작본(PROFILE_TEXT[lang])이 있으면 본문 채움, 없으면 placeholder."""
    label = cfg["label"]
    authored = PROFILE_TEXT.get(lang)

    def prof(lv: int, stage: str | None) -> str:
        if authored and lv in authored:
            return authored[lv]
        if lv == 1:
            return f"(T5 저작 예정) {label} 생존 표현 — 인사·숫자·정형표현."
        return f"(T5 저작 예정) {label} CEFR {stage} 레벨 발화 프로파일."

    levels = [{"level_no": 1, "band": "생존", "grade": None, "stage_name": STAGE_NAME[1],
               "textbook": None, "profile": prof(1, None)}]
    for stage in CEFR_STAGES:
        lv = STAGE_TO_LEVEL[stage]
        b = _band(lv)
        levels.append({
            "level_no": lv, "band": {1: "초급", 2: "중급", 3: "고급"}[b],
            "grade": {1: "A", 2: "B", 3: "C"}[b], "stage_name": STAGE_NAME[lv],
            "textbook": f"{lang.upper()}-{stage}", "profile": prof(lv, stage),
        })
    tag = "저작본" if authored else "골격"
    return {"_comment": f"{label} 레벨 프로파일({tag}, parse_lang). level_no 축 한국어와 동일(1=생존, 2~13=A1~C4).",
            "language": lang, "levels": levels}


def build_survival(lang: str) -> dict | None:
    """생존청크 46 → survival_chunks.json 페이로드. 저작본(SURVIVAL[lang]) 없으면 None."""
    raw = SURVIVAL.get(lang)
    if not raw:
        return None
    items = [
        {"no": i, "category": c, "surface": s, "reading": rd, "roman": rm,
         "meaning_en": en, "meaning_ko": ko, "situation": sit}
        for i, (c, s, rd, rm, en, ko, sit) in enumerate(raw, start=1)
    ]
    return {"level_no": 1, "count": len(items), "language": lang, "items": items}


def _strip(items):
    for it in items:
        for k in ("_si", "_no", "_seq"):
            it.pop(k, None)


def _dump(path: Path, payload):
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"  생성: {path.relative_to(_ROOT)}")


def run(lang: str) -> None:
    cfg = LANG_CONFIG[lang]
    print(f"\n=== [{lang}] {cfg['label']} ===")
    vocab = parse_vocab(lang, cfg)
    grammar = parse_grammar(lang, cfg)
    vdist = Counter(v["level_no"] for v in vocab)
    gdist = Counter(g["level_no"] for g in grammar)
    ko_hit = sum(1 for v in vocab if v.get("meanings"))
    print(f"  어휘 {len(vocab)} (한국어뜻 {ko_hit}={ko_hit/max(1,len(vocab)):.0%}) | 레벨분포 {dict(sorted(vdist.items()))}")
    print(f"  문법 {len(grammar)} | 레벨분포 {dict(sorted(gdist.items()))}")
    _strip(vocab); _strip(grammar)
    vocab.sort(key=lambda v: (v["level_no"], v["seq_no"]))
    grammar.sort(key=lambda g: (g["level_no"], g["seq_no"]))
    out = _ROOT / "assets" / "level" / f"curriculum_v2_{lang}"
    out.mkdir(parents=True, exist_ok=True)
    meta = {"source": f"level/05.다른 언어 CEFR/{cfg['folder']}", "generated_by": "scripts/curriculum/parse_lang.py", "language": lang}
    _dump(out / "grammar.json", {**meta, "count": len(grammar), "items": grammar})
    _dump(out / "vocab.json", {**meta, "count": len(vocab), "items": vocab})
    survival = build_survival(lang)
    if survival:
        _dump(out / "survival_chunks.json", survival)
        print(f"  생존청크 {survival['count']}개 (저작본)")
    _dump(_ROOT / "assets" / "level" / f"level_profiles_{lang}.json", build_profiles(lang, cfg))
    # 검증
    assert all(2 <= x["level_no"] <= 13 for x in vocab + grammar), "level_no 범위 이탈"
    assert len({v["source_key"] for v in vocab}) == len(vocab), "어휘 source_key 중복"
    assert len({g["source_key"] for g in grammar}) == len(grammar), "문법 source_key 중복"
    print("  ✓ 검증 통과")


def main() -> None:
    arg = sys.argv[1] if len(sys.argv) > 1 else "all"
    langs = list(LANG_CONFIG) if arg == "all" else [arg]
    for lang in langs:
        if lang not in LANG_CONFIG:
            raise SystemExit(f"미지원 언어: {lang} (지원: {list(LANG_CONFIG)} 또는 all)")
        run(lang)
    print("\n완료 ✅")


if __name__ == "__main__":
    main()
